#!/usr/bin/env python3
"""Combine the three watchlists into one tabbed HTML and one multi-sheet Excel file.

Page/Sheet A — VCP watchlist (latest, with tier trajectory across all snapshots)
Page/Sheet B — Weinstein Stage 2A watchlist (latest, with trajectory)
Page/Sheet C — Pre-breakout watchlist (latest, with trajectory)
Page/Sheet D — Summary of every ticker across all three lists (no history)

Usage: python make_bundle.py [--rev R0] [--model "Opus5;high"]
"""

from __future__ import annotations

import argparse
import collections
import glob
import html as html_mod
import json
import re
from datetime import datetime, timedelta, timezone

from exchanges import EXCHANGE, missing, tv_url

# ---------------------------------------------------------------- tier schemes
VCP_TIERS = {
    "A_VCP待突破": ("A", "VCP 緊縮・等待突破",
                    "趨勢模板通過、距 52 週高點 ≤10%、近 1 個月波動收斂 ±7% 內 — 最接近 Minervini VCP 買點。"),
    "E_突破延伸中": ("E", "已突破・延伸中",
                    "貼著 52 週新高且近 1 個月大漲 — 突破已發動，勿追高，等回測樞紐區或新的緊縮。"),
    "B_上升結構": ("B", "上升結構（一底高於一底 / 上升三角形候選）",
                  "站上 50/200 日線、距高點 ≤20%、3 個月動能為正 — 開圖確認低點墊高與上緣壓力線。"),
    "C_基底修復中": ("C", "基底修復中（觀察）", "仍在 200 日線上方但距高點較深，基底右側尚未完成。"),
    "D_趨勢弱": ("D", "趨勢偏弱（暫不列入）", "跌破主要均線或距高點過深，暫不符合進場條件。"),
}
STAGE_TIERS = {
    "2A_初升段": ("2A", "Stage 2A｜剛突破・初升段（首選）",
                  "站上 200 日（≈30 週）均線、距 52 週高點 ≤12%、升勢年輕（6 月 ≥+12%、1 年 ≤+100%）— Weinstein 最佳買進區。"),
    "2B_主升段": ("2B", "Stage 2B｜主升段・已延伸",
                  "升勢確立但漲幅已大 — 可持有，新買點等回測 30 週線。"),
    "1轉2_轉強觀察": ("1→2", "Stage 1→2｜轉強觀察",
                     "自第一階段基底翻揚初期，等待放量突破基底上緣確認。"),
    "3_做頭疑慮": ("3", "Stage 3｜做頭疑慮（線下）", "大漲後動能轉弱、距高點拉開 — 不宜新倉。"),
    "41_弱勢打底": ("4/1", "Stage 4／1｜下跌或打底（線下）", "跌勢未止或仍在第一階段基底中。"),
}
VCP_ONLINE = {"A", "E", "B"}
STAGE_ONLINE = {"2A", "2B", "1→2"}
VCP_ORDER = list(VCP_TIERS)
STAGE_ORDER = list(STAGE_TIERS)


def esc(s):
    return html_mod.escape(str(s), quote=True)


def off_cell(v):
    return "\u2013" if v is None else f"-{v:.1f}%"


def off_txt(r):
    v = r.get("off_high_pct")
    return "" if v is None else f" \u00b7 \u2212{v:.0f}%"


def num(v, dec=2):
    return "–" if v is None else f"{v:,.{dec}f}"


def pct(v):
    return "–" if v is None else f"{v:+.1f}%"


# ------------------------------------------------------------------ data load
def load_vcp_snapshots(limit=10):
    snaps = []
    for path in glob.glob("scan_R*.json"):
        m = re.match(r"scan_(R\d+)_", path)
        if not m:
            continue
        scan = json.load(open(path))
        rows = scan.get("rows", [])
        if not rows:
            continue
        d = collections.Counter((r.get("as_of") or "")[:10] for r in rows).most_common(1)[0][0]
        snaps.append((m.group(1), d, {r["ticker"]: r for r in rows}, scan))
    snaps.sort(key=lambda s: (s[1], int(s[0][1:])))
    return snaps[-limit:]


def load_single(path, key, rev="R0"):
    """A one-snapshot list: returns the same shape as load_vcp_snapshots."""
    scan = json.load(open(path))
    rows = scan["rows"]
    d = collections.Counter((r.get("as_of") or "")[:10] for r in rows).most_common(1)[0][0]
    for r in rows:
        if r.get(key) is not None:      # stage rows carry a stale VCP "category"
            r["category"] = r[key]
    return [(rev, d, {r["ticker"]: r for r in rows}, scan)]


def build_tracks(snaps, letter_of):
    tickers = sorted({t for _, _, rows, _ in snaps for t in rows})
    out = {}
    for t in tickers:
        letters, scores = [], []
        for _, _, rows, _ in snaps:
            r = rows.get(t)
            letters.append(letter_of.get(r["category"]) if r else None)
            scores.append(r.get("score") if r else None)
        out[t] = {"letters": letters, "scores": scores}
    return out


def rank_map(order, letter_of):
    return {letter_of[k]: i for i, k in enumerate(order)}


def move_of(info, ranks):
    seq = [l for l in info["letters"] if l]
    if len(seq) < 2:
        return 0
    return ranks[seq[-2]] - ranks[seq[-1]]


# ------------------------------------------------------------------ HTML page
def page_html(pid, title, subtitle, snaps, tracks, tiers, order, online, ranks, extra_cols):
    scan = snaps[-1][3]
    rows = scan["rows"]
    notes = {n["ticker"]: n for n in scan.get("notes", [])}
    letter_of = {k: v[0] for k, v in tiers.items()}
    n_snap = len(snaps)

    by = collections.defaultdict(list)
    dropped = []
    for r in rows:
        letter = letter_of[r["category"]]
        info = tracks.get(r["ticker"], {"letters": []})
        ever = any(l in online for l in info["letters"] if l)
        if letter in online:
            by[r["category"]].append(r)
        elif ever and n_snap > 1:
            dropped.append(r)
        else:
            by[r["category"]].append(r)
    counts = {c: len(by.get(c, [])) for c in order}
    n_online = sum(counts[c] for c in order if letter_of[c] in online)

    snap_heads = "".join(
        f'<th class="st">{r}<small>{d[5:].replace("-", "/")}</small></th>' for r, d, _, _ in snaps)
    head_cells = "".join(
        (f'<th>{c[0]}</th>' if c[2] == "txt" else f'<th class="num">{c[0]}</th>') for c in extra_cols)
    thead = (f'<thead><tr><th>代號</th><th>名稱</th>{snap_heads}<th class="tr">變化</th>'
             f'<th class="num">收盤</th>{head_cells}<th class="num">分數</th><th>備註</th></tr></thead>')

    def row_html(r):
        t = r["ticker"]
        info = tracks.get(t, {"letters": [None] * n_snap, "scores": [None] * n_snap})
        cells = ""
        for j, l in enumerate(info["letters"]):
            s = info["scores"][j]
            if not l:
                cells += '<td class="st none">–</td>'
            else:
                cls = f't{l.replace("→", "").replace("/", "")}'
                sc = "" if s is None else f"<i>{s:g}</i>"
                cells += f'<td class="st {cls}"><b>{l}</b>{sc}</td>'
        mv = move_of(info, ranks)
        arrow = ('<span class="mv up">▲</span>' if mv > 0 else
                 '<span class="mv dn">▼</span>' if mv < 0 else '<span class="mv fl">＝</span>')
        tds = ""
        for _, key, kind in extra_cols:
            v = r.get(key)
            if kind == "pct":
                cls = "" if v is None else ("up" if v > 0 else "dn" if v < 0 else "")
                tds += f'<td class="num {cls}">{pct(v)}</td>'
            elif kind == "offhigh":
                tds += f'<td class="num pivot">{"–" if v is None else f"-{v:.1f}%"}</td>'
            elif kind == "bool":
                tds += ('<td class="num up">✓</td>' if v else
                        '<td class="num dn">✗</td>' if v is False else '<td class="num">–</td>')
            elif kind == "vol":
                cls = "dry" if isinstance(v, (int, float)) and v < 0.7 else ""
                tds += f'<td class="num {cls}">{"–" if v is None else v}</td>'
            elif kind == "txt":
                tds += f'<td class="nm">{esc((v or "")[:26])}</td>'
            else:
                tds += f'<td class="num">{num(v)}</td>'
        note = esc(notes.get(t, {}).get("note", ""))
        ed = notes.get(t, {}).get("earnings_date", "")
        if ed:
            note += f' <span class="chip-cal">財報 {esc(ed[5:].replace("-", "/"))}</span>'
        ex = EXCHANGE.get(t, "").upper()
        score = r.get("score") or 0
        return (f'<tr><td class="tk"><a href="{tv_url(t)}" target="_blank" rel="noopener">{t}'
                f'<small>{ex}</small></a></td>'
                f'<td class="nm">{esc((r.get("name") or "")[:22])}</td>{cells}'
                f'<td class="tr">{arrow}</td><td class="num">{num(r.get("price"))}</td>{tds}'
                f'<td class="num"><span class="scorebar"><i style="width:{min(score, 100):.0f}%"></i>'
                f'<b>{score:g}</b></span></td><td class="note">{note}</td></tr>')

    def table(items):
        return (f'<div class="tblwrap"><table>{thead}<tbody>'
                + "\n".join(row_html(r) for r in items) + "</tbody></table></div>")

    secs = []
    for cat in order:
        letter = letter_of[cat]
        if letter not in online:
            continue
        items = sorted(by.get(cat, []), key=lambda r: -(r.get("score") or 0))
        if not items:
            continue
        tier, ttl, desc = tiers[cat]
        tid = tier.lower().replace("→", "").replace("/", "")
        secs.append(f'<section class="tier" id="{pid}-{tid}">'
                    f'<header class="tier-head"><span class="badge b-{tid}">{tier}</span>'
                    f'<div><h3>▲ 線上｜{esc(ttl)}</h3><p>{esc(desc)}</p></div>'
                    f'<span class="count">{len(items)} 檔</span></header>{table(items)}</section>')

    if dropped:
        secs.append('<div class="divider"><hr><span>以下為線下</span><hr></div>'
                    f'<section class="tier" id="{pid}-drop">'
                    '<header class="tier-head"><span class="badge b-drop">▼</span>'
                    '<div><h3>線下｜曾入選後跌出</h3><p>先前掃描曾列於線上級別，目前已跌出 — '
                    '若基底重新收緊可望回歸。</p></div>'
                    f'<span class="count">{len(dropped)} 檔</span></header>'
                    f'{table(sorted(dropped, key=lambda r: -(r.get("score") or 0)))}</section>')

    first_offline = True
    for cat in order:
        letter = letter_of[cat]
        if letter in online:
            continue
        items = sorted(by.get(cat, []), key=lambda r: -(r.get("score") or 0))
        if not items:
            continue
        tier, ttl, desc = tiers[cat]
        tid = tier.lower().replace("→", "").replace("/", "")
        if first_offline and not dropped:
            secs.append('<div class="divider"><hr><span>以下為線下</span><hr></div>')
        first_offline = False
        if len(items) > 40:
            chips = "".join(
                f'<a class="dchip" href="{tv_url(r["ticker"])}" target="_blank" rel="noopener">'
                f'{r["ticker"]}<small>{"".join(l or "–" for l in tracks[r["ticker"]]["letters"])}'
                f'{off_txt(r)}</small></a>'
                for r in sorted(items, key=lambda r: r.get("off_high_pct") or 0))
            body = f'<div class="dwrap">{chips}</div>'
        else:
            body = table(items)
        secs.append(f'<section class="tier" id="{pid}-{tid}">'
                    f'<header class="tier-head"><span class="badge b-{tid}">{tier}</span>'
                    f'<div><h3>▽ 線下｜{esc(ttl)}</h3><p>{esc(desc)}</p></div>'
                    f'<span class="count">{len(items)} 檔</span></header>{body}</section>')

    summary_rows = ""
    for rev_, d, srows, _ in snaps:
        c = collections.Counter(letter_of[r["category"]] for r in srows.values())
        cur = ' class="cur"' if rev_ == snaps[-1][0] else ""
        letters = [v[0] for v in tiers.values()]
        summary_rows += (f'<tr{cur}><td class="tk">{rev_}</td><td class="nm">{d}</td>'
                         + "".join(f'<td class="num">{c[l]}</td>' for l in letters)
                         + f'<td class="num">{sum(c.values())}</td></tr>')
    letters = [v[0] for v in tiers.values()]
    hist_head = "".join(f'<th class="num">{l}</th>' for l in letters)
    hist_note = ("" if n_snap > 1 else
                 '<p class="lede">本清單為首版，軌跡自 R0 起累積 — 下次更新後即可看出級別變化。</p>')

    stats = "".join(
        f'<a href="#{pid}-{v[0].lower().replace("→", "").replace("/", "")}">'
        f'<span class="stat"><span class="dot d-{v[0].lower().replace("→", "").replace("/", "")}"></span>'
        f'<b>{counts[k]}</b><span>{v[0]}</span></span></a>'
        for k, v in tiers.items() if counts.get(k))

    return f"""<section class="page" id="page-{pid}">
<h2 class="ptitle">{esc(title)}</h2>
<p class="lede">{subtitle}</p>
<nav class="summary">{stats}
<span class="stat tot"><b>{n_online}</b><span>線上合計</span></span>
<span class="stat tot"><b>{len(rows)}</b><span>全宇宙</span></span></nav>
<h4 class="sec">各期掃描概況</h4>
<div class="tblwrap"><table style="min-width:460px"><thead><tr><th>版本</th><th>數據日期</th>
{hist_head}<th class="num">合計</th></tr></thead><tbody>{summary_rows}</tbody></table></div>
{hist_note}
<p class="market"><strong>市場背景</strong> — {esc(scan.get('market', ''))}</p>
{''.join(secs)}
</section>"""


# ----------------------------------------------------------------- summary page
def build_summary(vcp, stage, pre):
    """ticker -> combined record across the three lists."""
    letters = {
        "vcp": {k: v[0] for k, v in VCP_TIERS.items()},
        "stage": {k: v[0] for k, v in STAGE_TIERS.items()},
        "pre": {k: v[0] for k, v in VCP_TIERS.items()},
    }
    rec = {}
    for key, snaps in (("vcp", vcp), ("stage", stage), ("pre", pre)):
        rows = snaps[-1][3]["rows"]
        for r in rows:
            e = rec.setdefault(r["ticker"], {"ticker": r["ticker"], "name": "", "price": None,
                                             "off_high_pct": None, "sector": "", "vcp": None, "stage": None,
                                             "pre": None, "vcp_s": None, "stage_s": None, "pre_s": None})
            e[key] = letters[key][r["category"]]
            e[f"{key}_s"] = r.get("score")
            if not e["name"] and r.get("name"):
                e["name"] = r["name"]
            if not e.get("sector") and r.get("sector"):
                e["sector"] = r["sector"]
            if e["price"] is None and r.get("price"):
                e["price"] = r["price"]
            if e["off_high_pct"] is None and r.get("off_high_pct") is not None:
                e["off_high_pct"] = r["off_high_pct"]
    for e in rec.values():
        on = 0
        on += 1 if e["vcp"] in VCP_ONLINE else 0
        on += 1 if e["stage"] in STAGE_ONLINE else 0
        on += 1 if e["pre"] in VCP_ONLINE else 0
        e["online_count"] = on
        e["lists"] = sum(1 for k in ("vcp", "stage", "pre") if e[k])
        e["best_score"] = max([s for s in (e["vcp_s"], e["stage_s"], e["pre_s"]) if s is not None], default=0)
    return rec


def summary_html(rec):
    def cell(v, online_set):
        if not v:
            return '<td class="gr none">–</td>'
        cls = "on" if v in online_set else "off"
        return f'<td class="gr {cls}"><b>{v}</b></td>'

    rows = sorted(rec.values(), key=lambda e: (-e["online_count"], -e["lists"], -e["best_score"]))
    trs = ""
    for e in rows:
        t = e["ticker"]
        ex = EXCHANGE.get(t, "").upper()
        trs += (f'<tr data-on="{e["online_count"]}"><td class="tk">'
                f'<a href="{tv_url(t)}" target="_blank" rel="noopener">{t}<small>{ex}</small></a></td>'
                f'<td class="nm">{esc((e["name"] or "")[:26])}</td>'
                f'<td class="nm">{esc((e.get("sector") or "")[:20])}</td>'
                + cell(e["vcp"], VCP_ONLINE) + cell(e["stage"], STAGE_ONLINE) + cell(e["pre"], VCP_ONLINE)
                + f'<td class="num"><b>{e["online_count"]}</b></td>'
                f'<td class="num">{num(e["price"])}</td>'
                f'<td class="num pivot">{off_cell(e["off_high_pct"])}</td>'
                f'<td class="num">{e["best_score"]:g}</td></tr>')
    n3 = sum(1 for e in rows if e["online_count"] == 3)
    n2 = sum(1 for e in rows if e["online_count"] == 2)
    n1 = sum(1 for e in rows if e["online_count"] == 1)
    return f"""<section class="page" id="page-d">
<h2 class="ptitle">Page D｜總表：所有代號 × 三份清單</h2>
<p class="lede">合併三份清單的全部 <b>{len(rows)}</b> 檔代號，列出每檔在各清單中的等級。
<b>綠底</b>＝該清單的線上級別（VCP：A／E／B；Weinstein：2A／2B／1→2）；灰底＝線下；「–」＝未出現在該清單。
依「線上清單數」排序 — 三榜皆線上者最值得優先開圖。本頁不含歷史軌跡（軌跡見 Page A／B／C）。</p>
<nav class="summary">
<span class="stat"><span class="dot d-a"></span><b>{n3}</b><span>三榜皆線上</span></span>
<span class="stat"><span class="dot d-b"></span><b>{n2}</b><span>兩榜線上</span></span>
<span class="stat"><span class="dot d-c"></span><b>{n1}</b><span>一榜線上</span></span>
<span class="stat tot"><b>{len(rows)}</b><span>代號總數</span></span></nav>
<div class="tblwrap"><table><thead><tr><th>代號</th><th>名稱</th>
<th>產業</th><th class="gr">VCP</th><th class="gr">Weinstein</th><th class="gr">Pre-breakout</th>
<th class="num">線上數</th><th class="num">收盤</th><th class="num">距高</th><th class="num">最高分</th>
</tr></thead><tbody>{trs}</tbody></table></div>
</section>"""


# --------------------------------------------------------------------- Excel
def write_excel(path, vcp, stage, pre, rec, rev, model, stamp_txt):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    head_fill = PatternFill("solid", fgColor="1B2430")
    head_font = Font(color="FFFFFF", bold=True, size=10)
    on_fill = PatternFill("solid", fgColor="D7EFE4")
    off_fill = PatternFill("solid", fgColor="EDEFEA")
    link_font = Font(color="0B6E4F", underline="single", bold=True)

    def style_sheet(ws, ncols, widths):
        ws.freeze_panes = "C2"
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        for c in range(1, ncols + 1):
            cell = ws.cell(row=1, column=c)
            cell.fill, cell.font = head_fill, head_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}{ws.max_row}"

    def list_sheet(ws, snaps, tiers, online, extra_cols, tracks):
        letter_of = {k: v[0] for k, v in tiers.items()}
        snap_revs = [s[0] for s in snaps]
        scan = snaps[-1][3]
        notes = {n["ticker"]: n for n in scan.get("notes", [])}
        cols = (["代號", "名稱", "交易所", "等級", "線上"] + [f"軌跡 {r}" for r in snap_revs]
                + ["收盤"] + [c[0] for c in extra_cols] + ["分數", "TradingView", "備註"])
        ws.append(cols)
        rows = sorted(scan["rows"],
                      key=lambda r: (list(tiers).index(r["category"]), -(r.get("score") or 0)))
        for r in rows:
            t = r["ticker"]
            info = tracks.get(t, {"letters": [None] * len(snap_revs)})
            letter = letter_of[r["category"]]
            line = ([t, (r.get("name") or "")[:40], EXCHANGE.get(t, "").upper(), letter,
                     "線上" if letter in online else "線下"]
                    + [l or "" for l in info["letters"]] + [r.get("price")]
                    + [r.get(c[1]) for c in extra_cols]
                    + [r.get("score"), tv_url(t), notes.get(t, {}).get("note", "")])
            ws.append(line)
            row = ws.max_row
            c = ws.cell(row=row, column=4)
            c.fill = on_fill if letter in online else off_fill
            c.alignment = Alignment(horizontal="center")
            c.font = Font(bold=True)
            lc = ws.cell(row=row, column=len(cols) - 1)
            lc.hyperlink = tv_url(t)
            lc.value = "chart"
            lc.font = link_font
            ws.cell(row=row, column=1).font = Font(bold=True)
        widths = ([9, 26, 9, 7, 7] + [8] * len(snap_revs) + [10]
                  + [20 if c[2] == "txt" else 9 for c in extra_cols] + [8, 10, 60])
        style_sheet(ws, len(cols), widths)

    wsA = wb.active
    wsA.title = "A. VCP"
    list_sheet(wsA, vcp[0], VCP_TIERS, VCP_ONLINE, vcp[2], vcp[1])
    wsB = wb.create_sheet("B. Weinstein 2A")
    list_sheet(wsB, stage[0], STAGE_TIERS, STAGE_ONLINE, stage[2], stage[1])
    wsC = wb.create_sheet("C. Pre-breakout")
    list_sheet(wsC, pre[0], VCP_TIERS, VCP_ONLINE, pre[2], pre[1])

    wsD = wb.create_sheet("D. Summary")
    wsD.append(["代號", "名稱", "產業", "交易所", "VCP", "Weinstein", "Pre-breakout",
                "線上清單數", "出現清單數", "收盤", "距高%", "最高分", "TradingView"])
    for e in sorted(rec.values(), key=lambda e: (-e["online_count"], -e["lists"], -e["best_score"])):
        t = e["ticker"]
        wsD.append([t, (e["name"] or "")[:40], (e.get("sector") or "")[:30],
                    EXCHANGE.get(t, "").upper(),
                    e["vcp"] or "", e["stage"] or "", e["pre"] or "",
                    e["online_count"], e["lists"], e["price"], e["off_high_pct"],
                    e["best_score"], tv_url(t)])
        row = wsD.max_row
        for col, key, onset in ((5, "vcp", VCP_ONLINE), (6, "stage", STAGE_ONLINE), (7, "pre", VCP_ONLINE)):
            c = wsD.cell(row=row, column=col)
            c.alignment = Alignment(horizontal="center")
            if e[key]:
                c.fill = on_fill if e[key] in onset else off_fill
                c.font = Font(bold=True)
        lc = wsD.cell(row=row, column=13)
        lc.hyperlink = tv_url(t)
        lc.value = "chart"
        lc.font = link_font
        wsD.cell(row=row, column=1).font = Font(bold=True)
    style_sheet(wsD, 13, [9, 28, 20, 9, 7, 10, 13, 10, 10, 10, 9, 8, 11])

    ws0 = wb.create_sheet("說明", 0)
    for line in [
        [f"Combined Watchlist {rev}"],
        [f"產生時間：{stamp_txt}｜模型：{model}"],
        [],
        ["工作表", "內容"],
        ["A. VCP", "Minervini VCP 分級清單（含各期軌跡欄）。線上＝A／E／B。"],
        ["B. Weinstein 2A", "Weinstein 階段分析（含軌跡欄）。線上＝2A／2B／1→2。"],
        ["C. Pre-breakout", "突破前候選清單（含軌跡欄）。線上＝A／E／B。"],
        ["D. Summary", "所有代號 × 三份清單的等級對照，不含歷史軌跡。"],
        [],
        ["等級底色", "綠＝該清單的線上級別；灰＝線下；空白＝未出現在該清單"],
        ["TradingView 欄", "點 chart 開該檔圖表"],
        [],
        ["注意", "本表為技術面選股輔助，非投資建議；分級由量化規則產生，形態請開圖確認。"],
    ]:
        ws0.append(line)
    ws0.column_dimensions["A"].width = 18
    ws0.column_dimensions["B"].width = 80
    ws0["A1"].font = Font(bold=True, size=14)
    for c in ("A4", "B4"):
        ws0[c].fill = head_fill
        ws0[c].font = head_font
    wb.save(path)


# ---------------------------------------------------------------------- main
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--rev", default="R0")
    ap.add_argument("--model", default="Opus5;high")
    args = ap.parse_args()

    vcp_snaps = load_vcp_snapshots()
    stage_snaps = load_single("scan_stage_R0_2026-08-22.json", "stage")
    pre_snaps = load_single("scan_PB-R0_2026-08-23.json", "category")

    vcp_letter = {k: v[0] for k, v in VCP_TIERS.items()}
    stage_letter = {k: v[0] for k, v in STAGE_TIERS.items()}
    vcp_tracks = build_tracks(vcp_snaps, vcp_letter)
    stage_tracks = build_tracks(stage_snaps, stage_letter)
    pre_tracks = build_tracks(pre_snaps, vcp_letter)
    vcp_ranks = rank_map(VCP_ORDER, vcp_letter)
    stage_ranks = rank_map(STAGE_ORDER, stage_letter)

    VCP_COLS = [("距高", "off_high_pct", "offhigh"), ("1月", "chg_1m", "pct"),
                ("3月", "chg_3m", "pct"), ("量比", "vol_ratio", "vol")]
    STAGE_COLS = [("距高", "off_high_pct", "offhigh"), ("6月", "chg_6m", "pct"),
                  ("1年", "chg_1y", "pct"), ("200日線上", "above_ma200", "bool")]
    PRE_COLS = [("產業", "sector", "txt"), ("距高", "off_high_pct", "offhigh"),
                ("1月", "chg_1m", "pct"), ("3月", "chg_3m", "pct"), ("量比", "vol_ratio", "vol")]

    all_missing = missing({r["ticker"] for s in (vcp_snaps, stage_snaps, pre_snaps)
                           for r in s[-1][3]["rows"]})
    if all_missing:
        print("WARNING unmapped exchanges:", all_missing)

    rec = build_summary(vcp_snaps, stage_snaps, pre_snaps)

    now_utc = datetime.now(timezone.utc)
    now_hkt = now_utc + timedelta(hours=8)
    stamp = now_hkt.strftime("%m.%d_%H.%M")
    stamp_txt = now_hkt.strftime("%Y.%m.%d %H:%M") + " HKT"
    base = f"Combined-VCP-2A-breakout(Git)_{args.rev} ({args.model})_({stamp})"

    pages = [
        page_html("a", "Page A｜VCP Watchlist（含歷史軌跡）",
                  f"Minervini 波動收縮形態。追蹤 {len(vcp_snaps)} 次更新，每列內含各期級別與分數。"
                  "線上＝A（緊縮待突破）／E（已突破延伸）／B（上升結構）。",
                  vcp_snaps, vcp_tracks, VCP_TIERS, VCP_ORDER, VCP_ONLINE, vcp_ranks, VCP_COLS),
        page_html("b", "Page B｜Weinstein Stage 2A Watchlist（含歷史軌跡）",
                  "Stan Weinstein 階段分析，重點在剛脫離第一階段基底的年輕升勢。"
                  "線上＝2A（初升段）／2B（主升段）／1→2（轉強觀察）。",
                  stage_snaps, stage_tracks, STAGE_TIERS, STAGE_ORDER, STAGE_ONLINE, stage_ranks, STAGE_COLS),
        page_html("c", "Page C｜Pre-Breakout Watchlist（含歷史軌跡）",
                  "突破前候選：貼近樞紐、量縮待變的標的。線上＝A／E／B。",
                  pre_snaps, pre_tracks, VCP_TIERS, VCP_ORDER, VCP_ONLINE, vcp_ranks, PRE_COLS),
        summary_html(rec),
    ]

    doc = HTML_SHELL.format(
        rev=args.rev, model=esc(args.model), stamp=stamp_txt,
        n_tickers=len(rec), pages="\n".join(pages),
        n_vcp=len(vcp_snaps[-1][3]["rows"]),
        n_stage=len(stage_snaps[-1][3]["rows"]),
        n_pre=len(pre_snaps[-1][3]["rows"]))
    open(f"{base}.html", "w").write(doc)

    write_excel(f"{base}.xlsx",
                (vcp_snaps, vcp_tracks, VCP_COLS),
                (stage_snaps, stage_tracks, STAGE_COLS),
                (pre_snaps, pre_tracks, PRE_COLS),
                rec, args.rev, args.model, stamp_txt)

    print(f"{base}.html")
    print(f"{base}.xlsx")
    print(f"tickers: {len(rec)} | vcp {len(vcp_snaps[-1][3]['rows'])} "
          f"| stage {len(stage_snaps[-1][3]['rows'])} | pre {len(pre_snaps[-1][3]['rows'])}")


HTML_SHELL = """<title>Combined Watchlist {rev}</title>
<style>
:root {{
  --bg:#F5F6F4; --panel:#FFFFFF; --ink:#1B2430; --muted:#5D6B7A; --line:#DDE2E0;
  --accent:#B07B24; --accent-soft:#F3E8D3; --up:#17835C; --dn:#C24A3F; --dry:#B07B24;
  --head:#EDEFEA; --hover:#F0F3EE; --on:#17835C; --off:#8B95A0;
  --tA:#17835C; --tE:#B07B24; --tB:#2C6E9E; --tC:#7A8794; --tD:#A8B0B8;
  --t2a:#17835C; --t2b:#2C6E9E; --t12:#B07B24; --t3:#A05A2C; --t41:#8B95A0; --tdrop:#C24A3F;
}}
@media (prefers-color-scheme: dark) {{
  :root:not([data-theme="light"]) {{
    --bg:#0E1216; --panel:#151B21; --ink:#E4E9EC; --muted:#8B98A5; --line:#26303A;
    --accent:#E5B15C; --accent-soft:#2A2418; --up:#3FB68B; --dn:#E0705F; --dry:#E5B15C;
    --head:#1A222A; --hover:#1C242D; --on:#3FB68B; --off:#5D6B7A;
    --tA:#3FB68B; --tE:#E5B15C; --tB:#5CA3D6; --tC:#6B7885; --tD:#4A555F;
    --t2a:#3FB68B; --t2b:#5CA3D6; --t12:#E5B15C; --t3:#D08A5A; --t41:#5D6B7A; --tdrop:#E0705F;
  }}
}}
:root[data-theme="dark"] {{
  --bg:#0E1216; --panel:#151B21; --ink:#E4E9EC; --muted:#8B98A5; --line:#26303A;
  --accent:#E5B15C; --accent-soft:#2A2418; --up:#3FB68B; --dn:#E0705F; --dry:#E5B15C;
  --head:#1A222A; --hover:#1C242D; --on:#3FB68B; --off:#5D6B7A;
  --tA:#3FB68B; --tE:#E5B15C; --tB:#5CA3D6; --tC:#6B7885; --tD:#4A555F;
  --t2a:#3FB68B; --t2b:#5CA3D6; --t12:#E5B15C; --t3:#D08A5A; --t41:#5D6B7A; --tdrop:#E0705F;
}}
* {{ box-sizing:border-box; }}
body {{ margin:0; background:var(--bg); color:var(--ink);
  font:15px/1.55 "Avenir Next","Segoe UI","PingFang TC","Microsoft JhengHei",system-ui,sans-serif; }}
.wrap {{ max-width:1380px; margin:0 auto; padding:26px 20px 60px; }}
.masthead {{ display:flex; flex-wrap:wrap; align-items:baseline; gap:10px 18px;
  border-bottom:3px solid var(--ink); padding-bottom:13px; }}
.masthead h1 {{ font-size:27px; margin:0; letter-spacing:-0.01em; }}
.masthead h1 em {{ font-style:normal; color:var(--accent); }}
.meta {{ color:var(--muted); font-size:13px; }} .meta b {{ color:var(--ink); font-weight:600; }}
.tabs {{ display:flex; flex-wrap:wrap; gap:6px; margin:18px 0 6px;
  border-bottom:1px solid var(--line); padding-bottom:0; }}
.tabs button {{ font:inherit; font-size:14px; font-weight:600; color:var(--muted); background:none;
  border:0; border-bottom:3px solid transparent; padding:9px 15px; cursor:pointer; border-radius:5px 5px 0 0; }}
.tabs button small {{ display:block; font-weight:400; font-size:11px; opacity:.75; }}
.tabs button:hover {{ background:var(--hover); color:var(--ink); }}
.tabs button[aria-selected="true"] {{ color:var(--accent); border-bottom-color:var(--accent); background:var(--panel); }}
.tabs button:focus-visible {{ outline:2px solid var(--accent); outline-offset:2px; }}
.page {{ display:none; }} .page.on {{ display:block; }}
.ptitle {{ font-size:21px; margin:22px 0 4px; }}
.lede {{ margin:6px 0 0; color:var(--muted); font-size:13.5px; max-width:84ch; }}
.lede b {{ color:var(--ink); }}
h4.sec {{ font-size:14px; margin:24px 0 8px; }}
.summary {{ display:flex; gap:9px; flex-wrap:wrap; margin:15px 0 6px; }}
.summary a {{ text-decoration:none; color:inherit; }}
.stat {{ display:flex; align-items:center; gap:8px; background:var(--panel);
  border:1px solid var(--line); border-radius:6px; padding:7px 13px; }}
.stat.tot {{ border-style:dashed; }}
.stat .dot {{ width:10px; height:10px; border-radius:2px; }}
.stat b {{ font-size:19px; font-variant-numeric:tabular-nums; }}
.stat span {{ color:var(--muted); font-size:12.5px; }}
.d-a, .d-2a {{ background:var(--tA); }} .d-e, .d-12 {{ background:var(--tE); }}
.d-b, .d-2b {{ background:var(--tB); }} .d-c, .d-3 {{ background:var(--tC); }}
.d-d, .d-41 {{ background:var(--tD); }}
.market {{ background:var(--panel); border:1px solid var(--line); border-left:4px solid var(--accent);
  border-radius:6px; padding:13px 17px; margin:14px 0 4px; font-size:13.5px; color:var(--muted); }}
.market strong {{ color:var(--ink); }}
.tier {{ margin-top:30px; }}
.tier-head {{ display:flex; align-items:flex-start; gap:13px; margin-bottom:9px; }}
.tier-head h3 {{ margin:0; font-size:17px; }}
.tier-head p {{ margin:2px 0 0; color:var(--muted); font-size:13px; max-width:74ch; }}
.tier-head .count {{ margin-left:auto; color:var(--muted); font-size:13px; white-space:nowrap; padding-top:3px; }}
.badge {{ flex:none; min-width:32px; height:32px; border-radius:6px; display:grid; place-items:center;
  font-weight:700; font-size:14px; color:var(--bg); padding:0 5px; }}
.b-a, .b-2a {{ background:var(--tA); }} .b-e, .b-12 {{ background:var(--tE); }}
.b-b, .b-2b {{ background:var(--tB); }} .b-c, .b-3 {{ background:var(--tC); }}
.b-d, .b-41 {{ background:var(--tD); }} .b-drop {{ background:var(--tdrop); }}
.divider {{ display:flex; align-items:center; gap:13px; margin:40px 0 0; }}
.divider hr {{ flex:1; border:0; border-top:2px dashed var(--accent); margin:0; }}
.divider span {{ color:var(--accent); font-size:12.5px; font-weight:700; letter-spacing:0.08em; }}
.tblwrap {{ overflow-x:auto; border:1px solid var(--line); border-radius:8px; background:var(--panel); }}
table {{ border-collapse:collapse; width:100%; font-size:13.5px; }}
th {{ background:var(--head); text-align:left; padding:7px 9px; font-size:11.5px; text-transform:uppercase;
  letter-spacing:0.04em; color:var(--muted); white-space:nowrap; position:sticky; top:0; }}
th small {{ display:block; font-size:10px; letter-spacing:0; text-transform:none; font-weight:400; }}
td {{ padding:7px 9px; border-top:1px solid var(--line); vertical-align:top; }}
tbody tr:hover {{ background:var(--hover); }}
td.num, th.num {{ text-align:right; font-variant-numeric:tabular-nums;
  font-family:"SF Mono","Cascadia Mono",Consolas,ui-monospace,monospace; font-size:12.5px; white-space:nowrap; }}
td.tk {{ white-space:nowrap; }}
td.tk a {{ display:inline-block; font-weight:700; color:var(--accent); text-decoration:none;
  border-bottom:1px solid transparent; }}
td.tk a small {{ display:block; font-size:9.5px; font-weight:600; letter-spacing:0.06em; color:var(--muted); }}
td.tk a:hover, td.tk a:focus-visible {{ border-bottom-color:var(--accent); outline:none; }}
td.nm {{ color:var(--muted); font-size:12.5px; white-space:nowrap; }}
th.st, td.st, th.gr, td.gr {{ text-align:center; }}
td.st {{ white-space:nowrap; font-variant-numeric:tabular-nums; }}
td.st b {{ display:inline-block; min-width:19px; height:19px; line-height:19px; border-radius:4px;
  color:var(--bg); font-size:11px; padding:0 3px; }}
td.st i {{ display:block; font-style:normal; font-size:10px; color:var(--muted); margin-top:1px; }}
td.st.none, td.gr.none {{ color:var(--muted); }}
td.tA b, td.t2A b {{ background:var(--tA); }} td.tE b {{ background:var(--tE); }}
td.tB b, td.t2B b {{ background:var(--tB); }} td.tC b, td.t3 b {{ background:var(--tC); }}
td.tD b, td.t41 b {{ background:var(--tD); }} td.t12 b {{ background:var(--t12); }}
td.gr b {{ display:inline-block; min-width:26px; padding:1px 6px; border-radius:4px;
  font-size:11.5px; color:var(--bg); }}
td.gr.on b {{ background:var(--on); }} td.gr.off b {{ background:var(--off); }}
tr.cur td {{ background:var(--accent-soft); font-weight:600; }}
td.tr, th.tr {{ text-align:center; }}
.mv {{ font-size:13px; }} .mv.up {{ color:var(--up); }} .mv.dn {{ color:var(--dn); }} .mv.fl {{ color:var(--muted); }}
td.up {{ color:var(--up); }} td.dn {{ color:var(--dn); }}
td.pivot {{ color:var(--accent); font-weight:600; }} td.dry {{ color:var(--dry); font-weight:600; }}
td.note {{ min-width:180px; max-width:340px; color:var(--muted); font-size:12px; line-height:1.45; white-space:normal; }}
.chip-cal {{ display:inline-block; background:var(--accent-soft); color:var(--accent); border-radius:4px;
  padding:1px 6px; font-size:11px; font-weight:600; white-space:nowrap; }}
.scorebar {{ display:inline-flex; align-items:center; gap:6px; min-width:84px; }}
.scorebar i {{ display:block; height:5px; border-radius:3px; background:var(--accent);
  min-width:3px; max-width:52px; flex:none; }}
.scorebar b {{ font-size:12.5px; }}
.dwrap {{ display:flex; flex-wrap:wrap; gap:6px; }}
.dchip {{ display:inline-flex; flex-direction:column; background:var(--panel); border:1px solid var(--line);
  border-radius:5px; padding:4px 9px; text-decoration:none; color:var(--muted); font-size:12.5px; font-weight:700; }}
.dchip small {{ font-weight:400; font-size:10px; letter-spacing:0.04em; }}
.dchip:hover, .dchip:focus-visible {{ border-color:var(--accent); color:var(--ink); outline:none; }}
.method {{ margin-top:40px; border-top:1px solid var(--line); padding-top:18px;
  color:var(--muted); font-size:13px; max-width:84ch; }}
.method h3 {{ color:var(--ink); font-size:15px; margin:0 0 8px; }}
.method ul {{ padding-left:20px; }} .method li {{ margin:4px 0; }}
.disclaimer {{ margin-top:13px; font-size:12.5px; border-left:3px solid var(--dn); padding-left:12px; }}
a {{ color:var(--accent); }}
@media (max-width:640px) {{ .masthead h1 {{ font-size:22px; }} .tier-head .count {{ display:none; }} }}
</style>
<div class="wrap">
<header class="masthead">
  <h1>Combined Watchlist <em>{rev}</em></h1>
  <span class="meta"><b>VCP · Weinstein 2A · Pre-breakout</b>｜合計 {n_tickers} 檔代號｜
  數據基準 2026-08-21 收盤｜產生 {stamp}｜{model}</span>
</header>
<nav class="tabs" role="tablist">
  <button role="tab" aria-selected="true" aria-controls="page-a" data-p="a">Page A · VCP<small>{n_vcp} 檔 · 含軌跡</small></button>
  <button role="tab" aria-selected="false" aria-controls="page-b" data-p="b">Page B · Weinstein 2A<small>{n_stage} 檔 · 含軌跡</small></button>
  <button role="tab" aria-selected="false" aria-controls="page-c" data-p="c">Page C · Pre-breakout<small>{n_pre} 檔 · 含軌跡</small></button>
  <button role="tab" aria-selected="false" aria-controls="page-d" data-p="d">Page D · 總表<small>{n_tickers} 檔 · 三榜對照</small></button>
</nav>
{pages}
<footer class="method">
<h3>使用說明</h3>
<ul>
<li><b>Page A／B／C</b> 各為一份清單的最新版本，每列的軌跡欄顯示該檔在歷次掃描中的級別與分數（有色方塊＝該期在線上）；變化欄 ▲ 升級、▼ 降級、＝ 持平。</li>
<li><b>Page D</b> 為三份清單的總表，列出每檔在各清單的等級與「線上清單數」— 三榜皆線上者最值得優先開圖確認。</li>
<li>三份清單使用不同方法：VCP 看波動收縮的買點、Weinstein 看升勢處於第幾階段、Pre-breakout 看貼近樞紐的待變標的 — 交集即為多重確認。</li>
<li>點任一代號可開 TradingView 圖表（Weinstein 建議切換週線 + 30 週均線確認階段）。</li>
</ul>
<p class="disclaimer">本表為技術面選股輔助工具，非投資建議。分級由量化規則產生，形態最終以圖表確認為準。</p>
</footer>
</div>
<script>
(function () {{
  var tabs = Array.prototype.slice.call(document.querySelectorAll('.tabs button'));
  var pages = Array.prototype.slice.call(document.querySelectorAll('.page'));
  function show(id, push) {{
    tabs.forEach(function (b) {{ b.setAttribute('aria-selected', b.dataset.p === id ? 'true' : 'false'); }});
    pages.forEach(function (p) {{ p.classList.toggle('on', p.id === 'page-' + id); }});
    try {{ localStorage.setItem('cw-page', id); }} catch (e) {{}}
    if (push) {{ try {{ history.replaceState(null, '', '#page-' + id); }} catch (e) {{}} }}
  }}
  tabs.forEach(function (b) {{
    b.addEventListener('click', function () {{ show(b.dataset.p, true); window.scrollTo({{top: 0}}); }});
    b.addEventListener('keydown', function (e) {{
      var i = tabs.indexOf(b), n = null;
      if (e.key === 'ArrowRight') n = tabs[(i + 1) % tabs.length];
      if (e.key === 'ArrowLeft') n = tabs[(i - 1 + tabs.length) % tabs.length];
      if (n) {{ e.preventDefault(); n.focus(); show(n.dataset.p, true); }}
    }});
  }});
  var start = 'a';
  var h = (location.hash || '').replace('#page-', '');
  if (['a', 'b', 'c', 'd'].indexOf(h) >= 0) start = h;
  else {{ try {{ var s = localStorage.getItem('cw-page'); if (s) start = s; }} catch (e) {{}} }}
  show(start, false);
  document.addEventListener('click', function (e) {{
    var a = e.target.closest && e.target.closest('a[href^="#"]');
    if (!a) return;
    var t = document.querySelector(a.getAttribute('href'));
    if (!t) return;
    var pg = t.closest('.page');
    if (pg && !pg.classList.contains('on')) show(pg.id.replace('page-', ''), true);
  }});
}})();
</script>"""


if __name__ == "__main__":
    main()

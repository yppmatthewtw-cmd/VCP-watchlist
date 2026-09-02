#!/usr/bin/env python3
"""Momentum Top50: 10-page tabbed HTML + 11-sheet Excel, with every change since the
previous release highlighted in red (diff engine below).

Pages 1a/1b/1c  — VCP TOP 50 by score, split by market cap (big/mid/small).
Pages 2a/2b/2c  — Weinstein 2A TOP 50 by cap tier.
Pages 3a/3b/3c  — Pre-breakout TOP 50 by cap tier.
Page 4          — grand summary of every ticker: three grades, upside-readiness
                  score, market cap, and the 7-item certainty evidence from the
                  10MA session as separate sortable columns (click = sort desc).

Cap tiers: big >= $10B, mid $2-10B, small < $2B (official snapshot market caps).
Usage: python make_momentum.py --rev R0 --model "Fable5.1;ultracode" \
         --prev scan_R15_2026-09-01.json,scan_stage_R8_2026-09-01.json,scan_PB-R8_2026-09-01.json
"""

from __future__ import annotations

import argparse
import collections
import html as html_mod
import json
import re
from datetime import datetime, timedelta, timezone

from exchanges import EXCHANGE, missing, tv_url
from make_bundle import (VCP_TIERS, STAGE_TIERS, VCP_ONLINE, STAGE_ONLINE,
                         esc, num, pct,
                         load_vcp_snapshots, load_series, build_summary)


def off_cell(v):
    if v is None:
        return "–"
    return "0.0%" if abs(v) < 0.05 else f"-{v:.1f}%"

CERT = json.load(open("cert7_2026-09-01.json"))
CATALYST = json.load(open("catalysts.json"))

C7_COLS = [("break", "突破", "25%", "兩底間中繼高點突破（未破按進度×0.6）"),
           ("retr", "回升", "10%", "跌幅收復比例"),
           ("time", "守底", "15%", "最後低點守住天數／15（跌破×0.25）"),
           ("dv", "量縮", "15%", "跌日／漲日量能比在全巿場合格股（約 3,000 檔）中的百分位（低者佳）"),
           ("contr", "收縮", "10%", "回檔幅度遞減在全巿場 HL 結構股中的百分位（收縮者佳）"),
           ("rs", "RS", "10%", "21 日報酬減全巿場中位數後的百分位"),
           ("ma", "均線", "15%", "收盤>MA20＋MA20>MA50＋MA50 上揚")]

CAP_TIERS = [("a", "大型股", "市值 ≥ $10B", lambda m: m >= 10e9),
             ("b", "中型股", "市值 $2B–10B", lambda m: 2e9 <= m < 10e9),
             ("c", "小型股", "市值 < $2B", lambda m: 0 < m < 2e9)]

# quote-verification chatter that leaked into notes — not for the spotlight pages
_NOISE = re.compile(r"(?i)quote|dated|session range|sourced|per the|multiple sources|"
                    r"confirm|unavailable|stale|re-?check|undated|plausible|vs prior|"
                    r"vs (the )?Aug|prior close|reference|clustered|live quotes?|"
                    r"\brange\b|price \d|\bAug \d|\d{1,2}/\d{1,2} [-+]\d|matching|baseline|"
                    r"approx|prev(ious)? close|not explicit|conflicting|consistent with|derived|cited|"
                    r"\bfigure\b|\bprice\b|vs \d{1,2}/\d{1,2}|reported|labeled|snippet|source")


def clean_note(txt, limit=80):
    parts = [p.strip() for p in re.split(r"[。;；]", txt or "") if p.strip() and not _NOISE.search(p)]
    out = "。".join(parts)
    if len(out) <= limit:
        return out
    cut = out[:limit]
    m = max(cut.rfind("。"), cut.rfind(" "), cut.rfind("，"))
    return (cut[:m] if m > limit // 2 else cut).rstrip("，。 ") + "…"


NEWEST = "2026-09-01"


PREV = {}          # ticker -> {"vcp": row, "stage": row, "pre": row} from the previous release
PREV_ROWS = {}     # key -> previous rows in file order (stable tie-break for rank moves)
ALIAS = {"GPS": "GAP"}
PREV_RANK = {}     # previous rank_overlay
PREV_CERT = {}     # previous cert7
PREV_CAT = {}      # previous catalysts


def load_prev(paths, rank_path, cert_path, cat_path):
    for key, p in zip(("vcp", "stage", "pre"), paths):
        for r in json.load(open(p))["rows"]:
            if key == "stage" and r.get("stage") is not None:
                r["category"] = r["stage"]      # stage rows carry a stale VCP "category"
            r["ticker"] = ALIAS.get(r["ticker"], r["ticker"])
            PREV.setdefault(r["ticker"], {})[key] = r
            PREV_ROWS.setdefault(key, []).append(r)
    for tgt, p in ((PREV_RANK, rank_path), (PREV_CERT, cert_path), (PREV_CAT, cat_path)):
        try:
            tgt.update(json.load(open(p)))
        except FileNotFoundError:
            pass


def changed(new, old, tol=0.0):
    """True when a value differs from the previous release beyond tol."""
    if old is None and new is None:
        return False
    if old is None or new is None:
        return True
    if isinstance(new, (int, float)) and isinstance(old, (int, float)):
        return abs(new - old) > tol
    return new != old


def upd(cell_html, is_changed, title=""):
    """Wrap a cell's inner HTML in the red 'updated' treatment when it changed."""
    if not is_changed:
        return cell_html
    t = f' title="{esc(title)}"' if title else ""
    return f'<span class="upd"{t}>{cell_html}</span>'


def grade_html(letter, old_letter):
    """Grade badge; when the tier changed, the old letter is struck through in red."""
    if old_letter and old_letter != letter:
        return f'<s class="old">{old_letter}</s><b class="upd">{letter}</b>'
    return f'<b>{letter}</b>'


def day_tag(v):
    if v is None:
        return ""
    cls = "up" if v > 0 else "dn" if v < 0 else ""
    return f'<small class="day {cls}">{v:+.1f}%</small>'


def asof_badge(d):
    """Mark a row whose price is not from the newest session, so the mixed
    basis is visible per row instead of only in the masthead."""
    d = (d or "")[:10]
    if not d or d == NEWEST:
        return ""
    return f'<small class="asof">{d[5:].replace("-", "/")}</small>'


TIER_RANK = {"A": 0, "B": 1, "E": 2, "C": 3, "D": 4,
             "2A": 0, "2B": 1, "1→2": 2, "3": 3, "4/1": 4}


def top50(rows, letter_of, online, pred):
    pool = [r for r in rows if pred(r.get("mcap") or 0)]
    return sorted(pool, key=lambda r: (letter_of.get(r["category"], "?") not in online,
                                       TIER_RANK.get(letter_of.get(r["category"], "?"), 9),
                                       -(r.get("score") or 0)))[:50], len(pool)


_BARE_MOVE = re.compile(r"^\s*\d{1,2}/\d{1,2}\s*[-+][\d.]+%\s*$")


def cat_removed(cat_old, cat_now):
    """Previous chip disappeared and it was real news (not just a bare day-move)."""
    return bool(cat_old) and not cat_now and not _BARE_MOVE.match(cat_old.get("reason") or "")


def mcap_txt(m):
    if not m:
        return "–"
    return f"${m/1e12:.2f}T" if m >= 1e12 else f"${m/1e9:.1f}B"


def _cat_sign(c):
    if c.get("pts"):
        return 1 if c["pts"] > 0 else -1
    mv = c.get("move")
    return 1 if (mv or 0) > 0 else -1


def cat_chip(t, long=False):
    c = CATALYST.get(t)
    if not c:
        return ""
    sign = _cat_sign(c)
    cls = "hot" if sign > 0 else "warn"
    icon = ("🔥" if sign > 0 else "⚠") if c.get("kind") == "news" or c.get("pts") else ("▲" if sign > 0 else "▼")
    return f'<span class="cat {cls}">{icon} {esc(c["reason"])}</span>'


def hotbar(tickers, limit=None):
    """Prominent, concise banner of this week's news-driven catalysts on a page."""
    hits = [(t, CATALYST[t]) for t in tickers if t in CATALYST]
    if not hits:
        return ""
    hits.sort(key=lambda x: -abs(x[1]["pts"]))
    more = ""
    if limit and len(hits) > limit:
        more = f'<span class="hmore">…另 {len(hits) - limit} 檔見表內催化欄</span>'
        hits = hits[:limit]
    chips = "".join(
        f'<a class="hchip {"hot" if c["pts"] > 0 else "warn"}" href="{tv_url(t)}" target="_blank" '
        f'rel="noopener"><b>{t}</b> {esc(c["reason"])}</a>' for t, c in hits)
    return (f'<div class="hotbar"><span class="hlabel">🔥 本週熱點催化</span>'
            f'<span class="hmacro">{MACRO}</span>'
            f'{chips}{more}</div>')


# ------------------------------------------------------------- cap-tier pages
def cap_page(pid, list_no, list_name, snaps, tiers, online, extra_cols,
             tier_name, tier_desc, pred):
    scan = snaps[-1][3]
    notes = {n["ticker"]: n for n in scan.get("notes", [])}
    letter_of = {k: v[0] for k, v in tiers.items()}
    no_mcap = sum(1 for r in scan["rows"] if not r.get("mcap"))
    items, n_pool = top50(scan["rows"], letter_of, online, pred)
    pool = [None] * n_pool

    # previous release's TOP-50 for this page (same rule, previous rows in their file order)
    pkey = {"1": "vcp", "2": "stage", "3": "pre"}[list_no]
    prev_items, _ = top50(PREV_ROWS.get(pkey, []), letter_of, online, pred)
    prev_rank = {r["ticker"]: i for i, r in enumerate(prev_items, 1)}
    cur_set = {r["ticker"] for r in items}
    entered = [r["ticker"] for r in items if r["ticker"] not in prev_rank]
    left = [r["ticker"] for r in prev_items if r["ticker"] not in cur_set]
    up_g, down_g, brk_g = [], [], []
    for r in items:
        po = PREV.get(r["ticker"], {}).get(pkey)
        if po and po.get("category") != r["category"] and po["category"] in tiers:
            a, b = letter_of.get(po["category"], "?"), letter_of[r["category"]]
            lab = f"{r['ticker']} {a}→{b}"
            if b == "E":
                brk_g.append(lab)
            elif a == "E":
                (up_g if b == "A" else down_g).append(lab)
            elif list(tiers).index(r["category"]) < list(tiers).index(po["category"]):
                up_g.append(lab)
            else:
                down_g.append(lab)

    head_extra = "".join(f'<th class="num sort">{c[0]}</th>' for c in extra_cols)
    trs = ""
    divider_done = False
    ncols = 12 + len(extra_cols)
    for i, r in enumerate(items, 1):
        t = r["ticker"]
        letter = letter_of[r["category"]]
        on = letter in online
        if not on and not divider_done:
            trs += (f'<tr class="divrow"><td colspan="{ncols}">▽ 以下為線下級別（該級距線上不足 50 檔，以分數補足）</td></tr>')
            divider_done = True
        po = PREV.get(t, {}).get(pkey) or {}
        old_letter = letter_of.get(po.get("category")) if po else None
        is_new = t not in prev_rank
        rank_move = (prev_rank[t] - i) if t in prev_rank else None
        tid = letter.lower().replace("→", "").replace("/", "")
        cert = CERT.get(t, {}).get("cert")
        note = esc(clean_note(notes.get(t, {}).get("note", "")))
        ed = notes.get(t, {}).get("earnings_date", "")
        if ed:
            note += f' <span class="chip-cal">財報 {esc(ed[5:].replace("-", "/"))}</span>'
        tds = ""
        for _, key, kind in extra_cols:
            v = r.get(key)
            ov = po.get(key) if po else None
            dv = v if isinstance(v, (int, float)) else ""
            ch = changed(v, ov, 0.05 if kind in ("pct", "vol") else 0)
            if kind == "pct":
                cls = "" if v is None else ("up" if v > 0 else "dn" if v < 0 else "")
                tds += f'<td class="num {cls}" data-v="{dv}">{upd(pct(v), ch, f"前版 {pct(ov)}")}</td>'
            elif kind == "vol":
                cls = "dry" if isinstance(v, (int, float)) and v < 0.7 else ""
                tds += f'<td class="num {cls}" data-v="{dv}">{upd("–" if v is None else str(v), ch)}</td>'
            elif kind == "bool":
                tds += (f'<td class="num up" data-v="1">{upd("✓", ch)}</td>' if v else
                        f'<td class="num dn" data-v="0">{upd("✗", ch)}</td>' if v is False
                        else '<td class="num" data-v="">–</td>')
            else:
                tds += f'<td class="num" data-v="{dv}">{upd(num(v), ch)}</td>'
        oh = r.get("off_high_pct")
        ooh = po.get("off_high_pct") if po else None
        score = r.get("score") or 0
        oscore = po.get("score") if po else None
        ocert = PREV_CERT.get(t, {}).get("cert")
        cert_txt = "–" if cert is None else f"{cert:g}"
        ocert_txt = "–" if ocert is None else str(ocert)
        new_chip = '<span class="chip-new">NEW</span>' if is_new else ""
        rank_chip = ("" if rank_move in (None, 0) else
                     f'<small class="rk {"up" if rank_move > 0 else "dn"}">{"▲" if rank_move > 0 else "▼"}{abs(rank_move)}</small>')
        cat_now, cat_old = CATALYST.get(t), PREV_CAT.get(t)
        cat_changed = (cat_now or {}).get("reason") != (cat_old or {}).get("reason")
        cat_cell = upd(cat_chip(t), cat_changed and bool(cat_now), "前版 " + ((cat_old or {}).get("reason") or "無"))
        if cat_removed(cat_old, cat_now):
            cat_cell = f'<span class="upd" title="前版 {esc(cat_old.get("reason", ""))}">（催化已移除）</span>'
        row_cls = ' class="isnew"' if is_new else ""
        old_price = po.get("price") if po else None
        if old_price and r.get("_split") and r["_split"].startswith("3:2"):
            old_price = round(old_price * 2 / 3, 2)      # show the previous price on the post-split basis
        trs += (f'<tr{row_cls}><td class="num" data-v="{-i}">{i}{rank_chip}</td>'
                f'<td class="tk"><a href="{tv_url(t)}" target="_blank" rel="noopener">{t}'
                f'<small>{EXCHANGE.get(t, "").upper()}</small></a>{new_chip}</td>'
                f'<td class="nm">{esc((r.get("name") or "")[:22])}</td>'
                f'<td class="nm">{esc((r.get("sector") or "")[:14])}</td>'
                f'<td class="st t{tid}" data-v="{"1" if on else "0"}">{grade_html(letter, old_letter)}</td>'
                f'<td class="catcell">{cat_cell}</td>'
                f'<td class="num" data-v="{r.get("price") or ""}">'
                f'{upd(num(r.get("price")), changed(r.get("price"), old_price, 0.005), f"前版 {num(old_price)}")}'
                f'{day_tag(r.get("chg_1d"))}{asof_badge(r.get("as_of"))}</td>'
                f'<td class="num pivot" data-v="{-oh if oh is not None else ""}">{upd(off_cell(oh), changed(oh, ooh, 0.05), f"前版 {off_cell(ooh)}")}</td>'
                f'{tds}'
                f'<td class="num" data-v="{(r.get("mcap") or 0)/1e9:.2f}">{mcap_txt(r.get("mcap"))}</td>'
                f'<td class="num" data-v="{cert if cert is not None else ""}">'
                f'{upd(cert_txt, changed(cert, ocert, 3), "前版 " + ocert_txt)}</td>'
                f'<td class="num" data-v="{score}"><span class="scorebar">'
                f'<i style="width:{min(score, 100):.0f}%"></i>{upd(f"<b>{score:g}</b>", changed(score, oscore, 0.05), f"前版 {oscore}")}</span></td>'
                f'<td class="note">{note}</td></tr>')

    n_on = sum(1 for r in items if letter_of[r["category"]] in online)
    cap_note = f'（此級距共 {len(pool)} 檔，取分數最高 {len(items)} 檔）' if len(pool) > 50 else \
        f'（此級距僅 {len(pool)} 檔，全數列出）' if len(pool) < 50 else ""
    miss_note = f'；另有 {no_mcap} 檔無市值數據未入級距' if no_mcap else ""
    return f"""<section class="page" id="page-{pid}">
<h2 class="ptitle">Page {pid}｜{list_name} TOP 50・{tier_name}（{tier_desc}）</h2>
<p class="lede">{list_name}清單內{tier_name}（依官方 9/1 市值），先按級別（{"A → B → E" if list_no != "2" else "2A → 2B → 1→2"}）、再按該清單分數由高至低{cap_note}{miss_note}。
線上級別 {n_on} 檔；欄位標題可點擊重新排序（先降序）。</p>
{hotbar([r["ticker"] for r in items])}
{update_panel(entered, left, up_g, down_g, items, prev_rank, brk_g)}
<div class="tblwrap"><table class="sortable"><thead><tr>
<th class="num sort">#</th><th>代號</th><th>名稱</th><th>產業</th><th class="st sort">等級</th>
<th>催化</th><th class="num sort">收盤</th><th class="num sort">距高</th>{head_extra}
<th class="num sort">市值</th><th class="num sort">確定性</th><th class="num sort">分數</th><th>備註</th>
</tr></thead><tbody>{trs}</tbody></table></div>
</section>"""


def chips(ts, show=14):
    """Chip list; beyond `show` items the rest fold into an inline <details>."""
    def one(t):
        return f'<a class="uchip" href="{tv_url(t.split()[0])}" target="_blank" rel="noopener">{esc(t)}</a>'
    if not ts:
        return '<span class="none">無</span>'
    if len(ts) <= show:
        return "".join(one(t) for t in ts)
    return ("".join(one(t) for t in ts[:show]) +
            f'<details><summary>…其餘 {len(ts) - show} 檔</summary>{"".join(one(t) for t in ts[show:])}</details>')


def update_panel(entered, left, up_g, down_g, items, prev_rank, brk_g=()):
    """Red 'what changed since the previous release' box at the top of a page."""
    movers = sorted((r for r in items if r.get("chg_1d") is not None), key=lambda r: r["chg_1d"])
    dn = "、".join(f"{r['ticker']} {r['chg_1d']:+.1f}%" for r in movers[:5])
    up = "、".join(f"{r['ticker']} {r['chg_1d']:+.1f}%" for r in movers[-5:][::-1])
    return f"""<div class="updbox"><div class="uhead"><span class="ulabel">本版更新（紅色＝相對前一版有變動）</span>
<span class="ulegend"><span class="upd">紅字</span>＝數值已變（滑鼠停留顯示前版值）· <span class="chip-new">NEW</span>＝新進榜 ·
<s class="old">A</s><b class="upd">B</b>＝等級變動 · <small class="rk up">▲3</small>／<small class="rk dn">▼3</small>＝名次升降 · <small class="day dn">-1.2%</small>＝9/1 當日漲跌</span></div>
<div class="urow"><b>新進榜 {len(entered)}</b>{chips(entered)}</div>
<div class="urow"><b>跌出榜 {len(left)}</b>{chips(left)}</div>
<div class="urow"><b>等級上調 {len(up_g)}</b>{chips(up_g)}</div>
<div class="urow"><b>突破延伸→E {len(brk_g)}</b>{chips(list(brk_g))}</div>
<div class="urow"><b>等級下調 {len(down_g)}</b>{chips(down_g)}</div>
<div class="urow"><b>9/1 當日</b><span class="utxt">最強：{esc(up) or "–"}｜最弱：{esc(dn) or "–"}</span></div><div class="urow"><b>規則修訂</b><span class="utxt">本版依批判性檢視修訂了分級規則（1 月漲幅 &gt;15% 視為延伸而非基底；真實 MA50 取代代理值；Stage 3 先於 1→2 判定；2A 需完整 6 月／1 年數據；52 週高低點以官方序列補正）— 部分等級變動源於規則而非價格，詳見頁尾。</span></div></div>"""


# ------------------------------------------------------------- page 4 summary
def page4_update_panel(rows):
    vl = {"vcp": {k: v[0] for k, v in VCP_TIERS.items()}, "stage": {k: v[0] for k, v in STAGE_TIERS.items()}}
    ups, downs, brk, rise_up, rise_dn = [], [], [], [], []
    order = {"vcp": list(vl["vcp"].values()), "stage": list(vl["stage"].values()), "pre": list(vl["vcp"].values())}
    for e in rows:
        t = e["ticker"]; pv = PREV.get(t, {})
        for k, key in (("vcp", "category"), ("stage", "stage"), ("pre", "category")):
            og = vl["vcp" if k != "stage" else "stage"].get((pv.get(k) or {}).get(key))
            ng = e.get(k)
            if og and ng and og != ng:
                lab = f"{t} {k[:3].upper()} {og}→{ng}"
                if ng == "E":
                    brk.append(lab)                       # broke out of the base: neither up nor down
                elif og == "E":
                    (ups if ng == "A" else downs).append(lab)   # E→A = re-tightened near highs
                else:
                    (ups if order[k].index(ng) < order[k].index(og) else downs).append(lab)
        prs = PREV_RANK.get(t, {}).get("score")
        if prs is not None and abs(e.get("rise_score", 0) - prs) >= 8:
            (rise_up if e["rise_score"] > prs else rise_dn).append(f"{t} {prs:g}→{e['rise_score']:g}")
    return f"""<div class="updbox"><div class="uhead"><span class="ulabel">本版更新（紅色＝相對前一版有變動）</span>
<span class="ulegend"><span class="upd">紅字</span>＝數值已變（滑鼠停留顯示前版值）· <s class="old">A</s><b class="upd">B</b>＝等級變動 · 確定性分項 |Δ|≥3 才標紅</span></div>
<div class="urow"><b>等級上調 {len(ups)}</b>{chips(ups)}</div>
<div class="urow"><b>突破延伸→E {len(brk)}</b>{chips(brk)}</div>
<div class="urow"><b>等級下調 {len(downs)}</b>{chips(downs)}</div>
<div class="urow"><b>上升分數 +8↑ {len(rise_up)}</b>{chips(rise_up)}</div>
<div class="urow"><b>上升分數 −8↓ {len(rise_dn)}</b>{chips(rise_dn)}</div><div class="urow"><b>規則修訂</b><span class="utxt">本版依批判性檢視修訂了分級規則（1 月漲幅 &gt;15% 視為延伸而非基底；真實 MA50 取代代理值；Stage 3 先於 1→2 判定；2A 需完整 6 月／1 年數據；52 週高低點以官方序列補正）— 部分等級變動源於規則而非價格，詳見頁尾。</span></div></div>"""


def summary4_html(rec):
    def cell(v, online_set, old=None):
        if not v:
            return '<td class="gr none" data-v="0">–</td>'
        cls = "on" if v in online_set else "off"
        inner = f'<s class="old">{old}</s><b class="upd">{v}</b>' if (old and old != v) else f'<b>{v}</b>'
        return f'<td class="gr {cls}" data-v="{2 if cls == "on" else 1}">{inner}</td>'

    rows = sorted(rec.values(), key=lambda e: (-e.get("rise_score", 0), -e["online_count"], -e["best_score"]))
    trs = ""
    for e in rows:
        t = e["ticker"]
        c = CERT.get(t)
        c7 = c["c7"] if c else None
        pc = PREV_CERT.get(t, {}).get("cert")
        pc_txt = "–" if pc is None else str(pc)
        cv = c["cert"] if c else None
        cert_tds = (
            f'<td class="num certsum" data-v="{cv}">'
            f'<span class="scorebar sb2"><i style="width:{min(cv, 100):.0f}%"></i>'
            f'{upd(f"<b>{cv:g}</b>", changed(cv, pc, 3), "前版 " + pc_txt)}</span>'
            f'{"" if c["hl_ok"] else "<small>無HL結構・僅3項</small>"}</td>'
            if c else '<td class="num certsum" data-v="">–</td>')
        oc = PREV_CERT.get(t)
        oc7 = oc["c7"] if oc else {}
        for k, _, _, _ in C7_COLS:
            cert_tds += (f'<td class="num c7" data-v="{c7[k]}">{upd(f"{c7[k]:g}", changed(c7[k], oc7.get(k), 3), "前版 " + str(oc7.get(k, "–")))}</td>'
                         if c7 else '<td class="num c7" data-v="">–</td>')
        mc = e.get("mcap") or 0
        cap = "大" if mc >= 10e9 else "中" if mc >= 2e9 else "小" if mc > 0 else "–"
        oh = e["off_high_pct"]
        why = esc(e.get("rise_why", ""))
        pv = PREV.get(t, {})
        vl = {"vcp": {k: v[0] for k, v in VCP_TIERS.items()}, "stage": {k: v[0] for k, v in STAGE_TIERS.items()}}
        old_g = {"vcp": vl["vcp"].get((pv.get("vcp") or {}).get("category")),
                 "stage": vl["stage"].get((pv.get("stage") or {}).get("stage")),
                 "pre": vl["vcp"].get((pv.get("pre") or {}).get("category"))}
        prs = PREV_RANK.get(t, {}).get("score")
        rs = e.get("rise_score", 0)
        cat_now, cat_old = CATALYST.get(t), PREV_CAT.get(t)
        cat_changed = (cat_now or {}).get("reason") != (cat_old or {}).get("reason")
        cat_cell = upd(cat_chip(t), cat_changed and bool(cat_now), "前版 " + ((cat_old or {}).get("reason") or "無"))
        if cat_removed(cat_old, cat_now):
            cat_cell = f'<span class="upd" title="前版 {esc(cat_old.get("reason", ""))}">（催化已移除）</span>'
        old_price = next((pv[k].get("price") for k in ("vcp", "stage", "pre") if k in pv), None)
        trs += (f'<tr><td class="tk"><a href="{tv_url(t)}" target="_blank" rel="noopener">{t}'
                f'<small>{EXCHANGE.get(t, "").upper()}</small></a></td>'
                f'<td class="nm">{esc((e["name"] or "")[:24])}</td>'
                f'<td class="nm">{esc((e.get("sector") or "")[:14])}</td>'
                + cell(e["vcp"], VCP_ONLINE, old_g["vcp"]) + cell(e["stage"], STAGE_ONLINE, old_g["stage"]) + cell(e["pre"], VCP_ONLINE, old_g["pre"])
                + f'<td class="num" data-v="{rs}"><span class="scorebar">'
                f'<i style="width:{min(rs, 100):.0f}%"></i>'
                f'{upd(f"<b>{rs:g}</b>", changed(rs, prs, 0.05), f"前版 {prs}")}</span></td>'
                f'<td class="whycell">{cat_cell}<span class="why">{why}</span></td>'
                f'<td class="num" data-v="{e["online_count"]}"><b>{e["online_count"]}</b></td>'
                f'<td class="num" data-v="{e["price"] or ""}">{upd(num(e["price"]), changed(e["price"], old_price, 0.005), f"前版 {num(old_price)}")}'
                f'{day_tag(e.get("chg_1d"))}{asof_badge(e.get("as_of"))}</td>'
                f'<td class="num pivot" data-v="{-oh if oh is not None else ""}">{off_cell(oh)}</td>'
                f'<td class="num" data-v="{mc/1e9:.2f}">{mcap_txt(mc)}<small class="captier">{cap}</small></td>'
                + cert_tds + "</tr>")

    n3 = sum(1 for e in rows if e["online_count"] == 3)
    n2 = sum(1 for e in rows if e["online_count"] == 2)
    n_cert = sum(1 for e in rows if e["ticker"] in CERT)
    n_hl = sum(1 for e in rows if CERT.get(e["ticker"], {}).get("hl_ok"))
    c7_heads = "".join(
        f'<th class="num sort c7" title="{esc(tip)}（權重 {w}）">{name}<small>{w}</small></th>'
        for _, name, w, tip in C7_COLS)
    return f"""<section class="page" id="page-4">
<h2 class="ptitle">Page 4｜總表：所有代號 × 三榜等級 × 確定性證據 7 項量化</h2>
<p class="lede">全部 <b>{len(rows)}</b> 檔代號。預設依「上升就緒分數」由高至低；<b>任何數值欄的標題都可點擊
重新排序（第一下＝降序，再點＝升序）</b>，包括右側 7 個確定性分項欄。確定性證據取自 10MA 上升趨勢
session 的 7 項量化（0–100，加權合計＝確定性總分）：{n_cert}/{len(rows)} 檔有官方日線序列可計算，
其中 <b>{n_hl}</b> 檔具「一底高於一底」結構（無結構者突破／回升／守底／收縮 4 項記 0，僅計量縮／RS／均線）；
「–」＝無足夠序列數據（多為外國 ADR）。</p>
{hotbar([e["ticker"] for e in rows if e["online_count"] >= 1], limit=24)}
{page4_update_panel(rows)}
<nav class="summary">
<span class="stat"><span class="dot d-a"></span><b>{n3}</b><span>三榜皆線上</span></span>
<span class="stat"><span class="dot d-b"></span><b>{n2}</b><span>兩榜線上</span></span>
<span class="stat"><span class="dot d-e"></span><b>{n_hl}</b><span>具 HL 結構</span></span>
<span class="stat tot"><b>{len(rows)}</b><span>代號總數</span></span></nav>
<div class="tblwrap"><table class="sortable"><thead><tr><th>代號</th><th>名稱</th><th>產業</th>
<th class="gr sort">VCP</th><th class="gr sort">2A</th><th class="gr sort">突破前</th>
<th class="num sort">上升分數</th><th>催化／主因</th><th class="num sort">線上</th>
<th class="num sort">收盤</th><th class="num sort">距高</th><th class="num sort">市值</th>
<th class="num sort certsum" title="7 項加權合計">確定性<small>總分</small></th>{c7_heads}
</tr></thead><tbody>{trs}</tbody></table></div>
</section>"""


# --------------------------------------------------------------------- Excel
def write_excel(path, cap_defs, rec, rev, model, stamp_txt, basis_txt):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    head_fill = PatternFill("solid", fgColor="1B2430")
    head_font = Font(color="FFFFFF", bold=True, size=10)
    on_fill = PatternFill("solid", fgColor="D7EFE4")
    off_fill = PatternFill("solid", fgColor="EDEFEA")
    hot_font = Font(color="B07B24", bold=True, size=9)
    warn_font = Font(color="C24A3F", bold=True, size=9)
    upd_font = Font(color="C8102E", bold=True)          # red = changed since the previous release
    upd_fill = PatternFill("solid", fgColor="FDE7EA")
    link_font = Font(color="0B6E4F", underline="single", bold=True)

    def style_sheet(ws, ncols, widths):
        ws.freeze_panes = "C2"
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        for c in range(1, ncols + 1):
            cell = ws.cell(row=1, column=c)
            cell.fill, cell.font = head_fill, head_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}{ws.max_row}"

    def cap_sheet(ws, snaps, tiers, online, extra_cols, pred):
        letter_of = {k: v[0] for k, v in tiers.items()}
        scan = snaps[-1][3]
        notes = {n["ticker"]: n for n in scan.get("notes", [])}
        items, _ = top50(scan["rows"], letter_of, online, pred)
        pkey = {"1": "vcp", "2": "stage", "3": "pre"}[pid[0]]
        prev_items, _ = top50(PREV_ROWS.get(pkey, []), letter_of, online, pred)
        prev_rank = {r["ticker"]: i for i, r in enumerate(prev_items, 1)}
        cols = (["排名", "前版名次", "代號", "名稱", "產業", "等級", "前版等級", "線上", "催化", "收盤", "基準日", "9/1當日%", "低於高點%"]
                + [c[0] for c in extra_cols] + ["市值($B)", "確定性", "前版確定性", "分數", "前版分數", "TradingView", "備註"])
        ws.append(cols)
        for i, r in enumerate(items, 1):
            t = r["ticker"]
            letter = letter_of[r["category"]]
            po = PREV.get(t, {}).get(pkey) or {}
            old_letter = letter_of.get(po.get("category"), "") if po else ""
            cat = CATALYST.get(t)
            cert = CERT.get(t, {}).get("cert")
            ws.append([i, prev_rank.get(t, "NEW"), t, (r.get("name") or "")[:40], (r.get("sector") or "")[:24],
                       letter, old_letter, "線上" if letter in online else "線下",
                       (("🔥 " if cat["pts"] > 0 else "⚠ ") + cat["reason"]) if cat else "",
                       r.get("price"), (r.get("as_of") or "")[:10], r.get("chg_1d"), r.get("off_high_pct")]
                      + [r.get(c[1]) for c in extra_cols]
                      + [round((r.get("mcap") or 0) / 1e9, 2) or None, cert, PREV_CERT.get(t, {}).get("cert"),
                         r.get("score"), po.get("score"), tv_url(t), clean_note(notes.get(t, {}).get("note", ""))])
            row = ws.max_row
            c = ws.cell(row=row, column=6)
            c.fill = on_fill if letter in online else off_fill
            c.alignment = Alignment(horizontal="center")
            c.font = Font(bold=True)
            if t not in prev_rank:                       # new entrant: whole row tinted red
                for col in range(1, len(cols) + 1):
                    ws.cell(row=row, column=col).fill = upd_fill
                ws.cell(row=row, column=2).font = upd_font
            if old_letter and old_letter != letter:      # grade change: red grade cell
                c.font = upd_font
            for col, key, tol in ((10, "price", 0.005), (13, "off_high_pct", 0.05), (len(cols) - 3, "score", 0.05)):
                if changed(r.get(key), po.get(key) if po else None, tol):
                    ws.cell(row=row, column=col).font = upd_font
            if changed(cert, PREV_CERT.get(t, {}).get("cert"), 3):
                ws.cell(row=row, column=len(cols) - 5).font = upd_font
            if cat and (PREV_CAT.get(t) or {}).get("reason") != cat["reason"]:
                ws.cell(row=row, column=9).font = upd_font
            elif cat:
                ws.cell(row=row, column=9).font = hot_font if cat["pts"] > 0 else warn_font
            lc = ws.cell(row=row, column=len(cols) - 1)
            lc.hyperlink = tv_url(t)
            lc.value = "chart"
            lc.font = link_font
            ws.cell(row=row, column=3).font = Font(bold=True)
        style_sheet(ws, len(cols), [6, 8, 9, 26, 16, 7, 8, 7, 30, 10, 11, 9, 10] + [9] * len(extra_cols)
                    + [10, 9, 9, 8, 8, 10, 50])

    first = True
    for pid, sheet_name, snaps, tiers, online, extra_cols, pred in cap_defs:
        if first:
            ws = wb.active
            ws.title = sheet_name
            first = False
        else:
            ws = wb.create_sheet(sheet_name)
        cap_sheet(ws, snaps, tiers, online, extra_cols, pred)

    ws4 = wb.create_sheet("4. 總表")
    ws4.append(["代號", "名稱", "產業", "交易所", "上升分數", "前版上升分數", "催化", "主因", "VCP", "前版VCP",
                "Weinstein", "前版Wein", "Pre-breakout", "前版PB", "線上數", "收盤", "9/1當日%", "低於高點%", "市值($B)",
                "級距", "確定性總分", "前版確定性"]
               + [f"{name}({w})" for _, name, w, _ in C7_COLS] + ["HL結構", "TradingView"])
    vl = {"vcp": {k: v[0] for k, v in VCP_TIERS.items()}, "stage": {k: v[0] for k, v in STAGE_TIERS.items()}}
    for e in sorted(rec.values(), key=lambda e: (-e.get("rise_score", 0), -e["online_count"], -e["best_score"])):
        t = e["ticker"]
        c = CERT.get(t)
        cat = CATALYST.get(t)
        mc = e.get("mcap") or 0
        cap = "大" if mc >= 10e9 else "中" if mc >= 2e9 else "小" if mc > 0 else ""
        pv = PREV.get(t, {})
        old_g = {"vcp": vl["vcp"].get((pv.get("vcp") or {}).get("category"), ""),
                 "stage": vl["stage"].get((pv.get("stage") or {}).get("stage"), ""),
                 "pre": vl["vcp"].get((pv.get("pre") or {}).get("category"), "")}
        prs = PREV_RANK.get(t, {}).get("score")
        pc = PREV_CERT.get(t, {}).get("cert")
        ws4.append([t, (e["name"] or "")[:40], (e.get("sector") or "")[:24],
                    EXCHANGE.get(t, "").upper(), e.get("rise_score", 0), prs,
                    (("🔥 " if cat["pts"] > 0 else "⚠ ") + cat["reason"]) if cat else "",
                    e.get("rise_why", ""), e["vcp"] or "", old_g["vcp"], e["stage"] or "", old_g["stage"],
                    e["pre"] or "", old_g["pre"], e["online_count"], e["price"], e.get("chg_1d"), e["off_high_pct"],
                    round(mc / 1e9, 2) or None, cap, c["cert"] if c else None, pc]
                   + [c["c7"][k] if c else None for k, _, _, _ in C7_COLS]
                   + [("是" if c["hl_ok"] else "否") if c else "", tv_url(t)])
        row = ws4.max_row
        for col, key, onset in ((9, "vcp", VCP_ONLINE), (11, "stage", STAGE_ONLINE), (13, "pre", VCP_ONLINE)):
            cc = ws4.cell(row=row, column=col)
            cc.alignment = Alignment(horizontal="center")
            if e[key]:
                cc.fill = on_fill if e[key] in onset else off_fill
                cc.font = upd_font if (old_g[key] and old_g[key] != e[key]) else Font(bold=True)
        if changed(e.get("rise_score", 0), prs, 0.05):
            ws4.cell(row=row, column=5).font = upd_font
        if c and changed(c["cert"], pc, 3):
            ws4.cell(row=row, column=21).font = upd_font
        if cat and (PREV_CAT.get(t) or {}).get("reason") != cat["reason"]:
            ws4.cell(row=row, column=7).font = upd_font
        elif cat:
            ws4.cell(row=row, column=7).font = hot_font if cat["pts"] > 0 else warn_font
        lc = ws4.cell(row=row, column=22 + len(C7_COLS) + 2)
        lc.hyperlink = tv_url(t)
        lc.value = "chart"
        lc.font = link_font
        ws4.cell(row=row, column=1).font = Font(bold=True)
    style_sheet(ws4, 22 + len(C7_COLS) + 2,
                [9, 26, 16, 9, 9, 9, 28, 30, 7, 7, 10, 8, 12, 8, 8, 10, 9, 9, 10, 6, 10, 9] + [8] * len(C7_COLS) + [8, 10])

    ws0 = wb.create_sheet("說明", 0)
    for line in [
        [f"Momentum Top50 {rev}"],
        [f"產生時間：{stamp_txt}｜模型：{model}｜數據基準：{basis_txt}"],
        [],
        ["工作表", "內容"],
        ["1a/1b/1c", "VCP TOP 50 大型（≥$10B）／中型（$2–10B）／小型（<$2B），按 VCP 分數排序。"],
        ["2a/2b/2c", "Weinstein 2A TOP 50，同樣按市值分三個級距。"],
        ["3a/3b/3c", "Pre-breakout TOP 50，按市值分三個級距。"],
        ["4. 總表", "全部代號 × 三榜等級 × 上升分數 × 確定性證據 7 項量化（各為獨立欄，可用篩選排序）。"],
        [],
        ["確定性 7 項", "突破25%＋回升10%＋守底15%＋量縮15%＋收縮10%＋RS10%＋均線15%（每項 0–100，加權＝總分）"],
        ["HL 結構", "「是」＝近 45 日存在一底高於一底結構；「否」＝突破／回升／守底／收縮 4 項記 0"],
        ["催化欄", "🔥＝本週正面新聞催化；⚠＝負面／風險事件。人工整理，隨新聞更新。"],
        ["紅色字／底色", "紅字＝相對前一版（Combined Watchlist R7）已變動的數值或等級；整列淡紅底＝本版新進榜；「前版…」欄為對照值。"],
        ["等級底色", "綠＝該清單的線上級別；灰＝線下；空白＝未出現在該清單"],
        [],
        ["注意", "本表為技術面選股輔助，非投資建議；分級由量化規則產生，形態請開圖確認。"],
    ]:
        ws0.append(line)
    ws0.column_dimensions["A"].width = 18
    ws0.column_dimensions["B"].width = 90
    ws0["A1"].font = Font(bold=True, size=14)
    for c in ("A4", "B4"):
        ws0[c].fill = head_fill
        ws0[c].font = head_font
    wb.save(path)


# ---------------------------------------------------------------------- main
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--rev", default="R3")
    ap.add_argument("--model", default="Fable5.1;ultracode")
    ap.add_argument("--prev", default="scan_R15_2026-09-01.json,scan_stage_R8_2026-09-01.json,scan_PB-R8_2026-09-01.json",
                    help="previous release's three snapshot files (vcp,stage,pre) for the red diff")
    ap.add_argument("--prev-rank", default="rank_overlay_R7.json")
    ap.add_argument("--prev-cert", default="cert7_2026-08-28.json")
    ap.add_argument("--prev-cat", default="catalysts_R7.json")
    args = ap.parse_args()
    load_prev(args.prev.split(","), args.prev_rank, args.prev_cert, args.prev_cat)

    vcp_snaps = load_vcp_snapshots()
    stage_snaps = load_series("scan_stage_R*.json", "stage")
    pre_snaps = load_series("scan_PB-R*.json", "category")

    # cap pages carry 距高 already; per-list momentum columns only
    CAP_VCP = [("1月", "chg_1m", "pct"), ("3月", "chg_3m", "pct"), ("量比", "vol_ratio", "vol")]
    CAP_STG = [("6月", "chg_6m", "pct"), ("1年", "chg_1y", "pct"), ("200日線上", "above_ma200", "bool")]
    CAP_PRE = [("1月", "chg_1m", "pct"), ("3月", "chg_3m", "pct"), ("量比", "vol_ratio", "vol")]

    all_missing = missing({r["ticker"] for s in (vcp_snaps, stage_snaps, pre_snaps)
                           for r in s[-1][3]["rows"]})
    if all_missing:
        print("WARNING unmapped exchanges:", all_missing)

    rec = build_summary(vcp_snaps, stage_snaps, pre_snaps)
    for key, snaps in (("vcp", vcp_snaps), ("stage", stage_snaps), ("pre", pre_snaps)):
        for r in snaps[-1][3]["rows"]:
            e = rec.get(r["ticker"])
            if e is not None and not e.get("mcap") and r.get("mcap"):
                e["mcap"] = r["mcap"]
            if e is not None and (r.get("as_of") or "") > (e.get("as_of") or ""):
                e["as_of"] = r.get("as_of")
            if e is not None and e.get("chg_1d") is None and r.get("chg_1d") is not None:
                e["chg_1d"] = r["chg_1d"]
            if e is not None and (r.get("as_of") or "") > (e.get("as_of") or ""):
                e["as_of"] = r.get("as_of")
            if e is not None and e.get("chg_1d") is None and r.get("chg_1d") is not None:
                e["chg_1d"] = r["chg_1d"]

    # count UNIQUE tickers, not row-instances across the three lists
    per_ticker = {}
    for s in (vcp_snaps, stage_snaps, pre_snaps):
        for r in s[-1][3]["rows"]:
            d = (r.get("as_of") or "")[:10]
            if d and d > per_ticker.get(r["ticker"], ""):
                per_ticker[r["ticker"]] = d
    dates = collections.Counter(per_ticker.values())
    n_all = len(per_ticker)
    newest = max(dates)
    n_newest = dates[newest]
    older = n_all - n_newest
    stale_names = sorted(t for t, d in per_ticker.items() if d != newest)
    basis = (f"{newest} 官方收盤（全部 {n_all} 檔）" if older == 0
             else f"{newest} 官方收盤（{n_newest}/{n_all} 檔；{'、'.join(stale_names)} 無資料源，沿用舊報價並以紅字標示基準日）"
             if older <= 3
             else f"混合基準 — {newest} 收盤僅 {n_newest}/{n_all} 檔，其餘 {older} 檔為較早收盤（各列價格下方紅字標示其基準日）")

    # macro line from the rows themselves (sector medians of the official 9/1 move)
    import statistics
    global MACRO
    by_sec, allmv = {}, []
    for sn in (vcp_snaps, stage_snaps, pre_snaps):
        for r in sn[-1][3]["rows"]:
            if r.get("chg_1d") is not None and r.get("sector"):
                by_sec.setdefault(r["sector"], {})[r["ticker"]] = r["chg_1d"]
    meds = sorted(((statistics.median(v.values()), k) for k, v in by_sec.items() if len(v) >= 4))
    for v in by_sec.values(): allmv += list(v.values())
    weakest = "、".join(f"{k} {m:+.1f}%" for m, k in meds[:2])
    strongest = "、".join(f"{k} {m:+.1f}%" for m, k in meds[-2:][::-1])
    MACRO = (f"宏觀（9/1 官方收盤）：美軍再度空襲伊朗革命衛隊目標（荷姆茲海峽），油價與債息同步走高；"
             f"標普 −0.3%、納指 −1.0%、道指 −0.8%。本清單 {len(allmv)} 檔中位數 {statistics.median(allmv):+.2f}%；"
             f"板塊中位數最弱：{weakest}；最強：{strongest}")

    now_hkt = datetime.now(timezone.utc) + timedelta(hours=8)
    stamp = now_hkt.strftime("%m.%d_%H.%M")
    stamp_txt = now_hkt.strftime("%Y.%m.%d %H:%M") + " HKT"
    base = f"Momentum_Top50 {args.rev} ({args.model})_({stamp})"

    lists3 = [("1", "VCP", vcp_snaps, VCP_TIERS, VCP_ONLINE, CAP_VCP),
              ("2", "Weinstein 2A", stage_snaps, STAGE_TIERS, STAGE_ONLINE, CAP_STG),
              ("3", "Pre-breakout", pre_snaps, VCP_TIERS, VCP_ONLINE, CAP_PRE)]

    pages = []
    cap_defs = []
    for no, lname, snaps, tiers, online, cols in lists3:
        for suffix, tname, tdesc, pred in CAP_TIERS:
            pid = f"{no}{suffix}"
            pages.append(cap_page(pid, no, lname, snaps, tiers, online, cols, tname, tdesc, pred))
            cap_defs.append((pid, f"{pid} {lname[:4]}{tname[:2]}", snaps, tiers, online, cols, pred))
    pages.append(summary4_html(rec))

    tab_defs = []
    for no, lname, *_ in lists3:
        for suffix, tname, _, _ in CAP_TIERS:
            short = {"1": "VCP", "2": "2A", "3": "突破前"}[no]
            tab_defs.append((f"{no}{suffix}", f"{no}{suffix} · {tname[:2]}", f"{short} TOP50"))
    tab_defs.append(("4", "4 · 總表", f"{len(rec)} 檔 · 7項確定性"))
    tabs = "".join(
        f'<button role="tab" aria-selected="{"true" if i == 0 else "false"}" '
        f'aria-controls="page-{p}" data-p="{p}">{lbl}<small>{sub}</small></button>'
        for i, (p, lbl, sub) in enumerate(tab_defs))
    page_ids = json.dumps([p for p, _, _ in tab_defs])

    doc = HTML_SHELL.format(
        rev=args.rev, model=esc(args.model), stamp=stamp_txt,
        n_tickers=len(rec), pages="\n".join(pages), basis=basis, tabs=tabs, page_ids=page_ids)
    open(f"{base}.html", "w").write(doc)

    write_excel(f"{base}.xlsx", cap_defs, rec, args.rev, args.model, stamp_txt, basis)

    print(f"{base}.html")
    print(f"{base}.xlsx")
    print(f"tickers: {len(rec)} | vcp {len(vcp_snaps[-1][3]['rows'])} "
          f"| stage {len(stage_snaps[-1][3]['rows'])} | pre {len(pre_snaps[-1][3]['rows'])}")


HTML_SHELL = """<title>Momentum Top50 {rev}</title>
<style>
:root {{
  color-scheme: dark;
  --bg:#0E1216; --panel:#151B21; --ink:#E4E9EC; --muted:#8B98A5; --line:#26303A;
  --accent:#E5B15C; --accent-soft:#2A2418; --up:#3FB68B; --dn:#E0705F; --dry:#E5B15C;
  --head:#1A222A; --hover:#1C242D; --on:#3FB68B; --off:#5D6B7A;
  --hot:#332812; --hotink:#F0C070; --warnbg:#381F1C; --warnink:#F0958A;
  --updink:#FF6B6B; --updbg:rgba(255,107,107,.10);
  --tA:#3FB68B; --tE:#E5B15C; --tB:#5CA3D6; --tC:#6B7885; --tD:#4A555F;
  --t2a:#3FB68B; --t2b:#5CA3D6; --t12:#E5B15C; --t3:#D08A5A; --t41:#5D6B7A; --tdrop:#E0705F;
}}
:root[data-theme="light"] {{
  --updink:#C8102E; --updbg:rgba(200,16,46,.07);
  color-scheme: light;
  --bg:#F5F6F4; --panel:#FFFFFF; --ink:#1B2430; --muted:#5D6B7A; --line:#DDE2E0;
  --accent:#B07B24; --accent-soft:#F3E8D3; --up:#17835C; --dn:#C24A3F; --dry:#B07B24;
  --head:#EDEFEA; --hover:#F0F3EE; --on:#17835C; --off:#8B95A0;
  --hot:#FBEEDD; --hotink:#8A5A10; --warnbg:#F9E3E0; --warnink:#9E3A30;
  --tA:#17835C; --tE:#B07B24; --tB:#2C6E9E; --tC:#7A8794; --tD:#A8B0B8;
  --t2a:#17835C; --t2b:#2C6E9E; --t12:#B07B24; --t3:#A05A2C; --t41:#8B95A0; --tdrop:#C24A3F;
}}
* {{ box-sizing:border-box; }}
body {{ margin:0; background:var(--bg); color:var(--ink);
  font:15px/1.55 "Avenir Next","Segoe UI","PingFang TC","Microsoft JhengHei",system-ui,sans-serif; }}
.wrap {{ max-width:1500px; margin:0 auto; padding:26px 20px 60px; }}
.masthead {{ display:flex; flex-wrap:wrap; align-items:baseline; gap:10px 18px;
  border-bottom:3px solid var(--ink); padding-bottom:13px; }}
.masthead h1 {{ font-size:27px; margin:0; letter-spacing:-0.01em; }}
.masthead h1 em {{ font-style:normal; color:var(--accent); }}
.meta {{ color:var(--muted); font-size:13px; flex-basis:100%; order:3; }}
.meta b {{ color:var(--ink); font-weight:600; }}
.themebtn {{ margin-left:auto; order:2; font:inherit; font-size:12.5px; font-weight:600; color:var(--muted);
  background:var(--panel); border:1px solid var(--line); border-radius:6px; padding:6px 12px;
  cursor:pointer; white-space:nowrap; align-self:center; }}
.themebtn:hover {{ color:var(--ink); border-color:var(--accent); }}
.masthead h1 {{ order:1; }}
.themebtn:focus-visible {{ outline:2px solid var(--accent); outline-offset:2px; }}
.tabs {{ display:flex; flex-wrap:wrap; gap:4px; margin:18px 0 6px;
  border-bottom:1px solid var(--line); padding-bottom:0; }}
.tabs button {{ font:inherit; font-size:13px; font-weight:600; color:var(--muted); background:none;
  border:0; border-bottom:3px solid transparent; padding:8px 10px; cursor:pointer; border-radius:5px 5px 0 0; }}
.tabs button small {{ display:block; font-weight:400; font-size:10.5px; opacity:.75; }}
.tabs button:hover {{ background:var(--hover); color:var(--ink); }}
.tabs button[aria-selected="true"] {{ color:var(--accent); border-bottom-color:var(--accent); background:var(--panel); }}
.tabs button:focus-visible {{ outline:2px solid var(--accent); outline-offset:2px; }}
.tabs button:nth-child(3), .tabs button:nth-child(6),
.tabs button:nth-child(9) {{ margin-right:12px; }}
.page {{ display:none; }} .page.on {{ display:block; }}
.ptitle {{ font-size:21px; margin:22px 0 4px; }}
.lede {{ margin:6px 0 0; color:var(--muted); font-size:13.5px; max-width:96ch; }}
.lede b {{ color:var(--ink); }}
h4.sec {{ font-size:14px; margin:24px 0 8px; }}
.hotbar {{ display:flex; flex-wrap:wrap; align-items:center; gap:7px; background:var(--panel);
  border:1px solid var(--line); border-left:4px solid var(--dn); border-radius:6px;
  padding:10px 14px; margin:14px 0 10px; }}
.hlabel {{ font-weight:800; font-size:13.5px; color:var(--dn); white-space:nowrap; }}
.hmacro {{ color:var(--muted); font-size:12px; flex-basis:100%; order:9; }}
.hchip {{ display:inline-block; border-radius:5px; padding:3px 9px; font-size:12px;
  text-decoration:none; white-space:nowrap; }}
.hchip b {{ font-weight:800; }}
.hchip.hot {{ background:var(--hot); color:var(--hotink); }}
.hchip.warn {{ background:var(--warnbg); color:var(--warnink); }}
.hchip:hover {{ filter:brightness(1.06); }}
.hmore {{ color:var(--muted); font-size:11.5px; }}
.cat {{ display:inline-block; border-radius:4px; padding:1px 7px; font-size:11px; font-weight:700;
  white-space:nowrap; }}
.cat.hot {{ background:var(--hot); color:var(--hotink); }}
.cat.warn {{ background:var(--warnbg); color:var(--warnink); }}
.catcell {{ max-width:190px; }} .whycell {{ max-width:250px; white-space:normal; }}
.whycell .why {{ display:block; color:var(--muted); font-size:11.5px; line-height:1.4; }}
.captier {{ display:inline-block; margin-left:5px; color:var(--muted); font-size:10px; }}
.summary {{ display:flex; gap:9px; flex-wrap:wrap; margin:15px 0 6px; }}
.summary a {{ text-decoration:none; color:inherit; }}
.stat {{ display:flex; align-items:center; gap:8px; background:var(--panel);
  border:1px solid var(--line); border-radius:6px; padding:7px 13px; }}
.stat.tot {{ border-style:dashed; }}
.stat .dot {{ width:10px; height:10px; border-radius:2px; }}
.stat b {{ font-size:19px; font-variant-numeric:tabular-nums; }}
.stat span {{ color:var(--muted); font-size:12.5px; }}
.d-a, .d-2a {{ background:var(--tA); }} .d-e, .d-12 {{ background:var(--tE); }}
.d-b, .d-2b {{ background:var(--tB); }} .d-c, .d-3 {{ background:var(--tC); }}
.d-d, .d-41 {{ background:var(--tD); }}
.market {{ background:var(--panel); border:1px solid var(--line); border-left:4px solid var(--accent);
  border-radius:6px; padding:13px 17px; margin:14px 0 4px; font-size:13.5px; color:var(--muted); }}
.market strong {{ color:var(--ink); }}
.tier {{ margin-top:30px; }}
.tier-head {{ display:flex; align-items:flex-start; gap:13px; margin-bottom:9px; }}
.tier-head h3 {{ margin:0; font-size:17px; }}
.tier-head p {{ margin:2px 0 0; color:var(--muted); font-size:13px; max-width:74ch; }}
.tier-head .count {{ margin-left:auto; color:var(--muted); font-size:13px; white-space:nowrap; padding-top:3px; }}
.badge {{ flex:none; min-width:32px; height:32px; border-radius:6px; display:grid; place-items:center;
  font-weight:700; font-size:14px; color:var(--bg); padding:0 5px; }}
.b-a, .b-2a {{ background:var(--tA); }} .b-e, .b-12 {{ background:var(--tE); }}
.b-b, .b-2b {{ background:var(--tB); }} .b-c, .b-3 {{ background:var(--tC); }}
.b-d, .b-41 {{ background:var(--tD); }} .b-drop {{ background:var(--tdrop); }}
.divider {{ display:flex; align-items:center; gap:13px; margin:40px 0 0; }}
.divider hr {{ flex:1; border:0; border-top:2px dashed var(--accent); margin:0; }}
.divider span {{ color:var(--accent); font-size:12.5px; font-weight:700; letter-spacing:0.08em; }}
.tblwrap {{ overflow-x:auto; border:1px solid var(--line); border-radius:8px; background:var(--panel); }}
table {{ border-collapse:collapse; width:100%; font-size:13.5px; }}
th {{ background:var(--head); text-align:left; padding:7px 9px; font-size:11.5px; text-transform:uppercase;
  letter-spacing:0.04em; color:var(--muted); white-space:nowrap; position:sticky; top:0; }}
th small {{ display:block; font-size:10px; letter-spacing:0; text-transform:none; font-weight:400; }}
th.sort {{ cursor:pointer; user-select:none; }}
th.sort:hover {{ color:var(--ink); background:var(--hover); }}
th.sort::after {{ content:"⇅"; opacity:.45; margin-left:4px; font-size:10px; }}
th.sort[data-dir="desc"]::after {{ content:"▼"; opacity:1; color:var(--accent); }}
th.sort[data-dir="asc"]::after {{ content:"▲"; opacity:1; color:var(--accent); }}
th.c7, td.c7 {{ border-left:1px dotted var(--line); }}
th.certsum, td.certsum {{ border-left:2px solid var(--accent); }}
td.certsum small {{ display:block; color:var(--muted); font-size:9.5px; }}
small.asof {{ display:block; font-size:9px; color:var(--dn); letter-spacing:.02em;
  font-weight:700; }}
.upd {{ color:var(--updink); font-weight:700; border-bottom:1px dotted var(--updink); cursor:help; }}
.upd b {{ color:var(--updink); }}
s.old {{ color:var(--muted); font-size:10px; margin-right:3px; text-decoration:line-through; }}
td.st b.upd, td.gr b.upd {{ background:var(--updink); color:#fff; border-bottom:0; box-shadow:0 0 0 2px var(--updbg); }}
.ulegend small.rk, .ulegend small.day {{ display:inline; }}
.ulegend s.old + b.upd {{ background:var(--updink); color:#fff; padding:0 5px; border-radius:3px; }}
.urow details {{ display:inline; }} .urow summary {{ display:inline; cursor:pointer; color:var(--updink);
  font-size:11.5px; font-weight:700; text-decoration:underline dotted; }}
.urow details[open] summary {{ display:none; }}
.chip-new {{ display:inline-block; margin-left:5px; background:var(--updink); color:#fff; font-size:9px;
  font-weight:800; letter-spacing:.06em; border-radius:3px; padding:1px 5px; vertical-align:middle; }}
tr.isnew td {{ background:var(--updbg); }}
tr.divrow td {{ background:var(--head); color:var(--accent); font-weight:700; font-size:12px;
  text-align:center; letter-spacing:.04em; border-top:2px dashed var(--accent); }}
small.rk {{ display:block; font-size:9px; font-weight:700; }} small.rk.up {{ color:var(--up); }} small.rk.dn {{ color:var(--dn); }}
small.day {{ display:block; font-size:9.5px; font-weight:600; }} small.day.up {{ color:var(--up); }} small.day.dn {{ color:var(--dn); }}
.updbox {{ border:1px solid var(--updink); border-left:4px solid var(--updink); background:var(--updbg);
  border-radius:6px; padding:10px 14px; margin:10px 0 12px; font-size:12.5px; }}
.uhead {{ display:flex; flex-wrap:wrap; gap:6px 16px; align-items:baseline; margin-bottom:6px; }}
.ulabel {{ font-weight:800; color:var(--updink); font-size:13.5px; }}
.ulegend {{ color:var(--muted); font-size:11.5px; }}
.urow {{ display:flex; flex-wrap:wrap; gap:5px; align-items:center; margin:3px 0; }}
.urow > b {{ color:var(--ink); min-width:110px; font-size:12px; }}
.urow .none {{ color:var(--muted); }} .urow .utxt {{ color:var(--ink); }}
.uchip {{ display:inline-block; background:var(--panel); border:1px solid var(--updink); color:var(--updink);
  border-radius:4px; padding:1px 7px; font-size:11.5px; font-weight:700; text-decoration:none; }}
.sb2 i {{ background:var(--tB) !important; }}
td {{ padding:7px 9px; border-top:1px solid var(--line); vertical-align:top; }}
tbody tr:hover {{ background:var(--hover); }}
td.num, th.num {{ text-align:right; font-variant-numeric:tabular-nums;
  font-family:"SF Mono","Cascadia Mono",Consolas,ui-monospace,monospace; font-size:12.5px; white-space:nowrap; }}
td.tk {{ white-space:nowrap; }}
td.tk a {{ display:inline-block; font-weight:700; color:var(--accent); text-decoration:none;
  border-bottom:1px solid transparent; }}
td.tk a small {{ display:block; font-size:9.5px; font-weight:600; letter-spacing:0.06em; color:var(--muted); }}
td.tk a:hover, td.tk a:focus-visible {{ border-bottom-color:var(--accent); outline:none; }}
td.nm {{ color:var(--muted); font-size:12.5px; white-space:nowrap; }}
th.st, td.st, th.gr, td.gr {{ text-align:center; }}
td.st {{ white-space:nowrap; font-variant-numeric:tabular-nums; }}
td.st b {{ display:inline-block; min-width:19px; height:19px; line-height:19px; border-radius:4px;
  color:var(--bg); font-size:11px; padding:0 3px; }}
td.st i {{ display:block; font-style:normal; font-size:10px; color:var(--muted); margin-top:1px; }}
td.st.none, td.gr.none {{ color:var(--muted); }}
td.tA b, td.t2A b, td.t2a b {{ background:var(--tA); }} td.tE b, td.te b {{ background:var(--tE); }}
td.tB b, td.t2B b, td.tb b, td.t2b b {{ background:var(--tB); }}
td.tC b, td.t3 b, td.tc b {{ background:var(--tC); }}
td.tD b, td.t41 b, td.td b {{ background:var(--tD); }} td.t12 b {{ background:var(--t12); }}
td.ta b {{ background:var(--tA); }}
td.gr b {{ display:inline-block; min-width:26px; padding:1px 6px; border-radius:4px;
  font-size:11.5px; color:var(--bg); }}
td.gr.on b {{ background:var(--on); }} td.gr.off b {{ background:var(--off); }}
tr.cur td {{ background:var(--accent-soft); font-weight:600; }}
td.tr, th.tr {{ text-align:center; }}
.mv {{ font-size:13px; }} .mv.up {{ color:var(--up); }} .mv.dn {{ color:var(--dn); }} .mv.fl {{ color:var(--muted); }}
td.up {{ color:var(--up); }} td.dn {{ color:var(--dn); }}
td.pivot {{ color:var(--accent); font-weight:600; }} td.dry {{ color:var(--dry); font-weight:600; }}
td.note {{ min-width:150px; max-width:300px; color:var(--muted); font-size:12px; line-height:1.45; white-space:normal; }}
.chip-cal {{ display:inline-block; background:var(--accent-soft); color:var(--accent); border-radius:4px;
  padding:1px 6px; font-size:11px; font-weight:600; white-space:nowrap; }}
.scorebar {{ display:inline-flex; align-items:center; gap:6px; min-width:84px; }}
.scorebar i {{ display:block; height:5px; border-radius:3px; background:var(--accent);
  min-width:3px; max-width:52px; flex:none; }}
.scorebar b {{ font-size:12.5px; }}
.dwrap {{ display:flex; flex-wrap:wrap; gap:6px; }}
.dchip {{ display:inline-flex; flex-direction:column; background:var(--panel); border:1px solid var(--line);
  border-radius:5px; padding:4px 9px; text-decoration:none; color:var(--muted); font-size:12.5px; font-weight:700; }}
.dchip small {{ font-weight:400; font-size:10px; letter-spacing:0.04em; }}
.dchip:hover, .dchip:focus-visible {{ border-color:var(--accent); color:var(--ink); outline:none; }}
.method {{ margin-top:40px; border-top:1px solid var(--line); padding-top:18px;
  color:var(--muted); font-size:13px; max-width:96ch; }}
.method h3 {{ color:var(--ink); font-size:15px; margin:0 0 8px; }}
.method ul {{ padding-left:20px; }} .method li {{ margin:4px 0; }}
.disclaimer {{ margin-top:13px; font-size:12.5px; border-left:3px solid var(--dn); padding-left:12px; }}
a {{ color:var(--accent); }}
@media (max-width:640px) {{ .masthead h1 {{ font-size:22px; }} .tier-head .count {{ display:none; }} }}
</style>
<div class="wrap">
<header class="masthead">
  <h1>Momentum Top50 <em>{rev}</em></h1>
  <span class="meta"><b>VCP · Weinstein 2A · Pre-breakout · 市值分級 · 確定性 7 項</b>｜合計 {n_tickers} 檔｜
  數據基準 {basis}｜產生 {stamp}｜{model}</span>
  <button class="themebtn" id="themebtn" type="button" aria-pressed="true">🌙 深色</button>
</header>
<nav class="tabs" role="tablist">{tabs}</nav>
{pages}
<footer class="method">
<h3>使用說明</h3>
<ul>
<li><b>Page 1a–3c</b>：每份清單依官方市值分三個級距（大 ≥$10B／中 $2–10B／小 <$2B），各取分數 TOP 50。頁頂紅框
「本週熱點催化」集中列出該頁涉及的新聞驅動事件（🔥 正面／⚠ 風險），表內催化欄逐檔標示。</li>
<li><b>Page 4 總表</b>：所有代號 × 三榜等級 × 上升就緒分數 × 市值 × <b>確定性證據 7 項量化</b>（突破／回升／守底／
量縮／收縮／RS／均線，各 0–100 獨立成欄）。<b>點任何數值欄標題即依該欄重新排序</b>（第一下降序、再點升序）。</li>
<li>確定性 7 項取自 10MA 上升趨勢清單的算法：以官方日線序列偵測「一底高於一底」結構後計算突破進度、跌幅收復、
低點守住天數、量能對比、回檔收縮、相對強度、均線排列；加權合計＝確定性總分（權重 25/10/15/15/10/10/15）。</li>
<li><b>數據基準</b>：本版為<b>完整的官方 2026-09-01 收盤重掃</b> —— 上游每日快照倉庫已恢復更新（9/1 22:44 UTC 收盤後快照），
全部 274 檔中 273 檔取得官方 9/1 收盤（僅 GPS 不在該資料源，沿用舊報價並以紅字標示基準日）。倉庫在 8/31 漏掉一天，
本版以 9/1 快照的 <code>price_change</code> 反推出每檔的官方 8/31 收盤補齊序列（該日成交量以前後兩日平均代替，
僅影響量縮分項）。1 月／3 月／5 日動能、MA50、市值皆重新自官方序列計算。</li>
<li><b>拆股調整</b>：CRWD 4:1（7/02）、KLAC 10:1（6/12）、DD 1:3 反向（6/24）及 <b>RUSHB／RUSHA 3:2（8/31 配發，已核對 8-K）</b>
均已調整；RUSHB 原掃描的 52 週高低點與均線已按 2/3 換算。確定性 7 項量化以延伸至 9/1 的官方序列重算（240/274 檔有完整序列，
49 檔具一底高於一底結構）。</li>
<li><b>紅色標示＝相對前一版（Combined Watchlist R7）的變動</b>：每頁頂部的「本版更新」框列出新進榜／跌出榜／等級升降／
當日最強最弱；表內紅字為數值已變（滑鼠停留顯示前版值）、<b>NEW</b> 為新進榜、劃線舊等級→紅色新等級為等級變動、
▲▼ 為名次升降、價格下方的小字為 9/1 當日漲跌。確定性分項 |Δ|≥3 才標紅，以免全表泛紅。</li>
<li><b>本版批判性檢視後的修訂</b>：(1) VCP — 21 日漲幅 &gt;15% 一律歸 E（延伸），不再被當作 A/B/C 基底；A 級另需 21 日收盤區間 ≤12%；
有真實 MA50 時一律採用（代理值只用於缺少的 MA200）。(2) Weinstein — Stage 3（做頭）先於 1→2 判定；動能改為「3 月 &gt;0 且 1 月 &gt;−8%」；
2A 需有完整 6 月／1 年數據且 1 月 ≤20%、6 月 ≤60%；2B 需漲幅已大或為 Stage 2 內的回測；1→2 上限距高 25%。(3) 52 週高低點以官方序列的
最高／最低收盤補正並在三份清單間統一，同一代號在各頁的距高一致。(4) 確定性的百分位改在全巿場合格股（約 3,000 檔，同 10MA 參考算法）內排名，
不再只在本清單 240 檔內排名。(5) 催化層：一日漲跌只作標示、不再計分（全表已同日定價，再計分屬重複）；只有新聞事件計分（±8 內）。
(6) GPS 更正為現行代號 GAP；DRH／HTO／LOB 交易所更正為 NYSE；RUSHB 3:2 拆股連同 6 月／1 年變動一併調整。</li>
<li><b>仍未解決（需你決定）</b>：本清單為 274 檔人工宇宙，只覆蓋 ≥$2B 美國本土股的約 12%（213／1,725 檔）；全巿場 5,517 檔的官方日線序列已在本地，
可改為資料驅動的全巿場掃描，但 52 週高低點與 MA200 只有 107 個交易日可算，需以「期間高低」代替 — 這是方法論的取捨，未擅自更動。
另 MA200 仍沿用原始掃描值（序列不足 200 日）。</li>
<li>點任一代號可開 TradingView 圖表（Weinstein 建議切週線＋30 週均線）。</li>
</ul>
<p class="disclaimer">本表為技術面選股輔助工具，非投資建議。分級由量化規則產生，催化欄為人工整理的本週事件，
形態最終以圖表確認為準。</p>
</footer>
</div>
<script>
(function () {{
  var root = document.documentElement, tb = document.getElementById('themebtn');
  function paint(mode) {{
    if (mode === 'light') root.setAttribute('data-theme', 'light');
    else root.removeAttribute('data-theme');
    tb.textContent = mode === 'light' ? '☀️ 淺色' : '🌙 深色';
    tb.setAttribute('aria-pressed', mode === 'light' ? 'false' : 'true');
  }}
  var saved = 'dark';
  try {{ saved = localStorage.getItem('cw3-theme') || 'dark'; }} catch (e) {{}}
  paint(saved);
  tb.addEventListener('click', function () {{
    var next = root.getAttribute('data-theme') === 'light' ? 'dark' : 'light';
    paint(next);
    try {{ localStorage.setItem('cw3-theme', next); }} catch (e) {{}}
  }});
}})();
(function () {{
  var tabs = Array.prototype.slice.call(document.querySelectorAll('.tabs button'));
  var pages = Array.prototype.slice.call(document.querySelectorAll('.page'));
  var ids = {page_ids};
  function show(id, push) {{
    tabs.forEach(function (b) {{ b.setAttribute('aria-selected', b.dataset.p === id ? 'true' : 'false'); }});
    pages.forEach(function (p) {{ p.classList.toggle('on', p.id === 'page-' + id); }});
    try {{ localStorage.setItem('cw3-page', id); }} catch (e) {{}}
    if (push) {{ try {{ history.replaceState(null, '', '#page-' + id); }} catch (e) {{}} }}
  }}
  tabs.forEach(function (b) {{
    b.addEventListener('click', function () {{ show(b.dataset.p, true); window.scrollTo({{top: 0}}); }});
    b.addEventListener('keydown', function (e) {{
      var i = tabs.indexOf(b), n = null;
      if (e.key === 'ArrowRight') n = tabs[(i + 1) % tabs.length];
      if (e.key === 'ArrowLeft') n = tabs[(i - 1 + tabs.length) % tabs.length];
      if (n) {{ e.preventDefault(); n.focus(); show(n.dataset.p, true); }}
    }});
  }});
  var start = '1a';
  var h = (location.hash || '').replace('#page-', '');
  if (ids.indexOf(h) >= 0) start = h;
  else {{ try {{ var s = localStorage.getItem('cw3-page'); if (s && ids.indexOf(s) >= 0) start = s; }} catch (e) {{}} }}
  show(start, false);
  document.addEventListener('click', function (e) {{
    var a = e.target.closest && e.target.closest('a[href^="#"]');
    if (!a) return;
    var t = document.querySelector(a.getAttribute('href'));
    if (!t) return;
    var pg = t.closest('.page');
    if (pg && !pg.classList.contains('on')) show(pg.id.replace('page-', ''), true);
  }});
  // column sorting: first click = descending, second = ascending
  document.querySelectorAll('table.sortable').forEach(function (tb) {{
    var ths = Array.prototype.slice.call(tb.querySelectorAll('thead th'));
    ths.forEach(function (th, ci) {{
      if (!th.classList.contains('sort')) return;
      th.addEventListener('click', function () {{
        var dir = th.dataset.dir === 'desc' ? 'asc' : 'desc';
        ths.forEach(function (o) {{ delete o.dataset.dir; }});
        th.dataset.dir = dir;
        var tbody = tb.tBodies[0];
        var rows = Array.prototype.slice.call(tbody.rows);
        rows.sort(function (a, b) {{
          var av = parseFloat(a.cells[ci].getAttribute('data-v'));
          var bv = parseFloat(b.cells[ci].getAttribute('data-v'));
          var an = isNaN(av), bn = isNaN(bv);
          if (an && bn) return 0;
          if (an) return 1;           // missing values always sink to the bottom
          if (bn) return -1;
          return dir === 'desc' ? bv - av : av - bv;
        }});
        rows.forEach(function (r) {{ tbody.appendChild(r); }});
      }});
    }});
  }});
}})();
</script>"""


if __name__ == "__main__":
    main()

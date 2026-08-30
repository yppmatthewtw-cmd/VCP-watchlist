#!/usr/bin/env python3
"""Combined Watchlist R3: 13-page tabbed HTML + 14-sheet Excel.

Pages A/B/C     — the three full watchlists with tier trajectories (as in R2.x).
Pages 1a/1b/1c  — VCP TOP 50 by score, split by market cap (big/mid/small).
Pages 2a/2b/2c  — Weinstein 2A TOP 50 by cap tier.
Pages 3a/3b/3c  — Pre-breakout TOP 50 by cap tier.
Page 4          — grand summary of every ticker: three grades, upside-readiness
                  score, market cap, and the 7-item certainty evidence from the
                  10MA session as separate sortable columns (click = sort desc).

Cap tiers: big >= $10B, mid $2-10B, small < $2B (official snapshot market caps).
Usage: python make_bundle3.py [--rev R3] [--model "Fable5;ultracode"]
"""

from __future__ import annotations

import argparse
import collections
import html as html_mod
import json
import re
from datetime import datetime, timedelta, timezone

from exchanges import EXCHANGE, missing, tv_url
from make_bundle import (VCP_TIERS, STAGE_TIERS, VCP_ONLINE, STAGE_ONLINE,
                         VCP_ORDER, STAGE_ORDER, esc, num, pct, off_cell,
                         load_vcp_snapshots, load_series, build_tracks,
                         rank_map, page_html, build_summary)

CERT = json.load(open("cert7_2026-08-28.json"))
CATALYST = json.load(open("catalysts.json"))

C7_COLS = [("break", "突破", "25%", "兩底間中繼高點突破（未破按進度×0.6）"),
           ("retr", "回升", "10%", "跌幅收復比例"),
           ("time", "守底", "15%", "最後低點守住天數／15（跌破×0.25）"),
           ("dv", "量縮", "15%", "跌日／漲日量能比的宇宙百分位（低者佳）"),
           ("contr", "收縮", "10%", "回檔幅度遞減的百分位（收縮者佳）"),
           ("rs", "RS", "10%", "21 日相對大盤中位數報酬百分位"),
           ("ma", "均線", "15%", "收盤>MA20＋MA20>MA50＋MA50 上揚")]

CAP_TIERS = [("a", "大型股", "市值 ≥ $10B", lambda m: m >= 10e9),
             ("b", "中型股", "市值 $2B–10B", lambda m: 2e9 <= m < 10e9),
             ("c", "小型股", "市值 < $2B", lambda m: 0 < m < 2e9)]

# quote-verification chatter that leaked into notes — not for the spotlight pages
_NOISE = re.compile(r"(?i)quote|dated|session range|sourced|per the|multiple sources|"
                    r"confirm|unavailable|stale|re-?check|undated|plausible|vs prior|"
                    r"vs (the )?Aug|prior close|reference|clustered|live quotes?")


def clean_note(txt, limit=70):
    parts = [p for p in re.split(r"[。;；]", txt or "") if p.strip() and not _NOISE.search(p)]
    return "。".join(parts)[:limit]


def mcap_txt(m):
    if not m:
        return "–"
    return f"${m/1e12:.2f}T" if m >= 1e12 else f"${m/1e9:.1f}B"


def cat_chip(t, long=False):
    c = CATALYST.get(t)
    if not c:
        return ""
    cls = "hot" if c["pts"] > 0 else "warn"
    icon = "🔥" if c["pts"] > 0 else "⚠"
    return f'<span class="cat {cls}">{icon} {esc(c["reason"])}</span>'


def hotbar(tickers):
    """Prominent, concise banner of this week's news-driven catalysts on a page."""
    hits = [(t, CATALYST[t]) for t in tickers if t in CATALYST]
    if not hits:
        return ""
    hits.sort(key=lambda x: -abs(x[1]["pts"]))
    chips = "".join(
        f'<a class="hchip {"hot" if c["pts"] > 0 else "warn"}" href="{tv_url(t)}" target="_blank" '
        f'rel="noopener"><b>{t}</b> {esc(c["reason"])}</a>' for t, c in hits)
    return (f'<div class="hotbar"><span class="hlabel">🔥 本週熱點催化</span>'
            f'<span class="hmacro">宏觀：Warsh 鷹派首秀，9 月加息機率約 55%，週五科技股回吐</span>'
            f'{chips}</div>')


# ------------------------------------------------------------- cap-tier pages
def cap_page(pid, list_no, list_name, snaps, tiers, online, extra_cols,
             tier_name, tier_desc, pred):
    scan = snaps[-1][3]
    notes = {n["ticker"]: n for n in scan.get("notes", [])}
    letter_of = {k: v[0] for k, v in tiers.items()}
    pool = [r for r in scan["rows"] if pred(r.get("mcap") or 0)]
    no_mcap = sum(1 for r in scan["rows"] if not r.get("mcap"))
    items = sorted(pool, key=lambda r: (letter_of[r["category"]] not in online,
                                        -(r.get("score") or 0)))[:50]

    head_extra = "".join(f'<th class="num sort">{c[0]}</th>' for c in extra_cols)
    trs = ""
    for i, r in enumerate(items, 1):
        t = r["ticker"]
        letter = letter_of[r["category"]]
        on = letter in online
        tid = letter.lower().replace("→", "").replace("/", "")
        cert = CERT.get(t, {}).get("cert")
        note = esc(clean_note(notes.get(t, {}).get("note", "")))
        ed = notes.get(t, {}).get("earnings_date", "")
        if ed:
            note += f' <span class="chip-cal">財報 {esc(ed[5:].replace("-", "/"))}</span>'
        tds = ""
        for _, key, kind in extra_cols:
            v = r.get(key)
            dv = v if isinstance(v, (int, float)) else ""
            if kind == "pct":
                cls = "" if v is None else ("up" if v > 0 else "dn" if v < 0 else "")
                tds += f'<td class="num {cls}" data-v="{dv}">{pct(v)}</td>'
            elif kind == "vol":
                cls = "dry" if isinstance(v, (int, float)) and v < 0.7 else ""
                tds += f'<td class="num {cls}" data-v="{dv}">{"–" if v is None else v}</td>'
            elif kind == "bool":
                tds += (f'<td class="num up" data-v="1">✓</td>' if v else
                        f'<td class="num dn" data-v="0">✗</td>' if v is False
                        else '<td class="num" data-v="">–</td>')
            else:
                tds += f'<td class="num" data-v="{dv}">{num(v)}</td>'
        oh = r.get("off_high_pct")
        score = r.get("score") or 0
        trs += (f'<tr><td class="num" data-v="{-i}">{i}</td>'
                f'<td class="tk"><a href="{tv_url(t)}" target="_blank" rel="noopener">{t}'
                f'<small>{EXCHANGE.get(t, "").upper()}</small></a></td>'
                f'<td class="nm">{esc((r.get("name") or "")[:22])}</td>'
                f'<td class="nm">{esc((r.get("sector") or "")[:14])}</td>'
                f'<td class="st t{tid}" data-v="{"1" if on else "0"}"><b>{letter}</b></td>'
                f'<td class="catcell">{cat_chip(t)}</td>'
                f'<td class="num" data-v="{r.get("price") or ""}">{num(r.get("price"))}</td>'
                f'<td class="num pivot" data-v="{-oh if oh is not None else ""}">{off_cell(oh)}</td>'
                f'{tds}'
                f'<td class="num" data-v="{(r.get("mcap") or 0)/1e9:.2f}">{mcap_txt(r.get("mcap"))}</td>'
                f'<td class="num" data-v="{cert if cert is not None else ""}">'
                f'{"–" if cert is None else f"{cert:g}"}</td>'
                f'<td class="num" data-v="{score}"><span class="scorebar">'
                f'<i style="width:{min(score, 100):.0f}%"></i><b>{score:g}</b></span></td>'
                f'<td class="note">{note}</td></tr>')

    n_on = sum(1 for r in items if letter_of[r["category"]] in online)
    cap_note = f'（此級距共 {len(pool)} 檔，取分數最高 {len(items)} 檔）' if len(pool) > 50 else \
        f'（此級距僅 {len(pool)} 檔，全數列出）' if len(pool) < 50 else ""
    miss_note = f'；另有 {no_mcap} 檔無市值數據未入級距' if no_mcap else ""
    return f"""<section class="page" id="page-{pid}">
<h2 class="ptitle">Page {pid}｜{list_name} TOP 50・{tier_name}（{tier_desc}）</h2>
<p class="lede">{list_name}清單內{tier_name}（依官方 8/28 市值），先列線上級別、再按該清單分數由高至低{cap_note}{miss_note}。
線上級別 {n_on} 檔；欄位標題可點擊重新排序（先降序）。</p>
{hotbar([r["ticker"] for r in items])}
<div class="tblwrap"><table class="sortable"><thead><tr>
<th class="num sort">#</th><th>代號</th><th>名稱</th><th>產業</th><th class="st sort">等級</th>
<th>催化</th><th class="num sort">收盤</th><th class="num sort">距高</th>{head_extra}
<th class="num sort">市值</th><th class="num sort">確定性</th><th class="num sort">分數</th><th>備註</th>
</tr></thead><tbody>{trs}</tbody></table></div>
</section>"""


# ------------------------------------------------------------- page 4 summary
def summary4_html(rec):
    def cell(v, online_set):
        if not v:
            return '<td class="gr none" data-v="0">–</td>'
        cls = "on" if v in online_set else "off"
        return f'<td class="gr {cls}" data-v="{2 if cls == "on" else 1}"><b>{v}</b></td>'

    rows = sorted(rec.values(), key=lambda e: (-e.get("rise_score", 0), -e["online_count"], -e["best_score"]))
    trs = ""
    for e in rows:
        t = e["ticker"]
        c = CERT.get(t)
        c7 = c["c7"] if c else None
        cert_tds = (
            f'<td class="num certsum" data-v="{c["cert"]}">'
            f'<span class="scorebar sb2"><i style="width:{min(c["cert"], 100):.0f}%"></i>'
            f'<b>{c["cert"]:g}</b></span>{"" if c["hl_ok"] else "<small>無HL結構</small>"}</td>'
            if c else '<td class="num certsum" data-v="">–</td>')
        for k, _, _, _ in C7_COLS:
            cert_tds += (f'<td class="num c7" data-v="{c7[k]}">{c7[k]:g}</td>'
                         if c7 else '<td class="num c7" data-v="">–</td>')
        mc = e.get("mcap") or 0
        cap = "大" if mc >= 10e9 else "中" if mc >= 2e9 else "小" if mc > 0 else "–"
        oh = e["off_high_pct"]
        why = esc(e.get("rise_why", ""))
        trs += (f'<tr><td class="tk"><a href="{tv_url(t)}" target="_blank" rel="noopener">{t}'
                f'<small>{EXCHANGE.get(t, "").upper()}</small></a></td>'
                f'<td class="nm">{esc((e["name"] or "")[:24])}</td>'
                f'<td class="nm">{esc((e.get("sector") or "")[:14])}</td>'
                + cell(e["vcp"], VCP_ONLINE) + cell(e["stage"], STAGE_ONLINE) + cell(e["pre"], VCP_ONLINE)
                + f'<td class="num" data-v="{e.get("rise_score", 0)}"><span class="scorebar">'
                f'<i style="width:{min(e.get("rise_score", 0), 100):.0f}%"></i>'
                f'<b>{e.get("rise_score", 0):g}</b></span></td>'
                f'<td class="whycell">{cat_chip(t)}<span class="why">{why}</span></td>'
                f'<td class="num" data-v="{e["online_count"]}"><b>{e["online_count"]}</b></td>'
                f'<td class="num" data-v="{e["price"] or ""}">{num(e["price"])}</td>'
                f'<td class="num pivot" data-v="{-oh if oh is not None else ""}">{off_cell(oh)}</td>'
                f'<td class="num" data-v="{mc/1e9:.2f}">{mcap_txt(mc)}<small class="captier">{cap}</small></td>'
                + cert_tds + "</tr>")

    n3 = sum(1 for e in rows if e["online_count"] == 3)
    n2 = sum(1 for e in rows if e["online_count"] == 2)
    n_cert = sum(1 for e in rows if e["ticker"] in CERT)
    n_hl = sum(1 for e in rows if CERT.get(e["ticker"], {}).get("hl_ok"))
    c7_heads = "".join(
        f'<th class="num sort c7" title="{esc(tip)}（權重 {w}）">{name}<small>{w}</small></th>'
        for _, name, w, tip in C7_COLS)
    return f"""<section class="page" id="page-4">
<h2 class="ptitle">Page 4｜總表：所有代號 × 三榜等級 × 確定性證據 7 項量化</h2>
<p class="lede">全部 <b>{len(rows)}</b> 檔代號。預設依「上升就緒分數」由高至低；<b>任何數值欄的標題都可點擊
重新排序（第一下＝降序，再點＝升序）</b>，包括右側 7 個確定性分項欄。確定性證據取自 10MA 上升趨勢
session 的 7 項量化（0–100，加權合計＝確定性總分）：{n_cert}/{len(rows)} 檔有官方日線序列可計算，
其中 <b>{n_hl}</b> 檔具「一底高於一底」結構（無結構者突破／回升／守底／收縮 4 項記 0，僅計量縮／RS／均線）；
「–」＝無足夠序列數據（多為外國 ADR）。</p>
{hotbar([e["ticker"] for e in rows if e["online_count"] >= 1])}
<nav class="summary">
<span class="stat"><span class="dot d-a"></span><b>{n3}</b><span>三榜皆線上</span></span>
<span class="stat"><span class="dot d-b"></span><b>{n2}</b><span>兩榜線上</span></span>
<span class="stat"><span class="dot d-e"></span><b>{n_hl}</b><span>具 HL 結構</span></span>
<span class="stat tot"><b>{len(rows)}</b><span>代號總數</span></span></nav>
<div class="tblwrap"><table class="sortable"><thead><tr><th>代號</th><th>名稱</th><th>產業</th>
<th class="gr sort">VCP</th><th class="gr sort">2A</th><th class="gr sort">突破前</th>
<th class="num sort">上升分數</th><th>催化／主因</th><th class="num sort">線上</th>
<th class="num sort">收盤</th><th class="num sort">距高</th><th class="num sort">市值</th>
<th class="num sort certsum" title="7 項加權合計">確定性<small>總分</small></th>{c7_heads}
</tr></thead><tbody>{trs}</tbody></table></div>
</section>"""


# --------------------------------------------------------------------- Excel
def write_excel(path, vcp, stage, pre, cap_defs, rec, rev, model, stamp_txt):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    head_fill = PatternFill("solid", fgColor="1B2430")
    head_font = Font(color="FFFFFF", bold=True, size=10)
    on_fill = PatternFill("solid", fgColor="D7EFE4")
    off_fill = PatternFill("solid", fgColor="EDEFEA")
    hot_font = Font(color="B07B24", bold=True, size=9)
    warn_font = Font(color="C24A3F", bold=True, size=9)
    link_font = Font(color="0B6E4F", underline="single", bold=True)

    def style_sheet(ws, ncols, widths):
        ws.freeze_panes = "C2"
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        for c in range(1, ncols + 1):
            cell = ws.cell(row=1, column=c)
            cell.fill, cell.font = head_fill, head_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}{ws.max_row}"

    def list_sheet(ws, snaps, tiers, online, extra_cols, tracks):
        letter_of = {k: v[0] for k, v in tiers.items()}
        snap_revs = [s[0] for s in snaps]
        scan = snaps[-1][3]
        notes = {n["ticker"]: n for n in scan.get("notes", [])}
        cols = (["代號", "名稱", "交易所", "等級", "線上"] + [f"軌跡 {r}" for r in snap_revs]
                + ["收盤"] + [c[0] for c in extra_cols] + ["分數", "TradingView", "備註"])
        ws.append(cols)
        rows = sorted(scan["rows"],
                      key=lambda r: (list(tiers).index(r["category"]), -(r.get("score") or 0)))
        for r in rows:
            t = r["ticker"]
            info = tracks.get(t, {"letters": [None] * len(snap_revs)})
            letter = letter_of[r["category"]]
            ws.append([t, (r.get("name") or "")[:40], EXCHANGE.get(t, "").upper(), letter,
                       "線上" if letter in online else "線下"]
                      + [l or "" for l in info["letters"]] + [r.get("price")]
                      + [r.get(c[1]) for c in extra_cols]
                      + [r.get("score"), tv_url(t), notes.get(t, {}).get("note", "")])
            row = ws.max_row
            c = ws.cell(row=row, column=4)
            c.fill = on_fill if letter in online else off_fill
            c.alignment = Alignment(horizontal="center")
            c.font = Font(bold=True)
            lc = ws.cell(row=row, column=len(cols) - 1)
            lc.hyperlink = tv_url(t)
            lc.value = "chart"
            lc.font = link_font
            ws.cell(row=row, column=1).font = Font(bold=True)
        style_sheet(ws, len(cols), ([9, 26, 9, 7, 7] + [8] * len(snap_revs) + [10]
                                    + [20 if c[2] == "txt" else 9 for c in extra_cols] + [8, 10, 60]))

    def cap_sheet(ws, snaps, tiers, online, extra_cols, pred):
        letter_of = {k: v[0] for k, v in tiers.items()}
        scan = snaps[-1][3]
        notes = {n["ticker"]: n for n in scan.get("notes", [])}
        pool = [r for r in scan["rows"] if pred(r.get("mcap") or 0)]
        items = sorted(pool, key=lambda r: (letter_of[r["category"]] not in online,
                                            -(r.get("score") or 0)))[:50]
        cols = (["排名", "代號", "名稱", "產業", "等級", "線上", "催化", "收盤", "距高%"]
                + [c[0] for c in extra_cols] + ["市值($B)", "確定性", "分數", "TradingView", "備註"])
        ws.append(cols)
        for i, r in enumerate(items, 1):
            t = r["ticker"]
            letter = letter_of[r["category"]]
            cat = CATALYST.get(t)
            cert = CERT.get(t, {}).get("cert")
            ws.append([i, t, (r.get("name") or "")[:40], (r.get("sector") or "")[:24], letter,
                       "線上" if letter in online else "線下",
                       (("🔥 " if cat["pts"] > 0 else "⚠ ") + cat["reason"]) if cat else "",
                       r.get("price"), r.get("off_high_pct")]
                      + [r.get(c[1]) for c in extra_cols]
                      + [round((r.get("mcap") or 0) / 1e9, 2) or None, cert,
                         r.get("score"), tv_url(t), clean_note(notes.get(t, {}).get("note", ""), 80)])
            row = ws.max_row
            c = ws.cell(row=row, column=5)
            c.fill = on_fill if letter in online else off_fill
            c.alignment = Alignment(horizontal="center")
            c.font = Font(bold=True)
            if cat:
                ws.cell(row=row, column=7).font = hot_font if cat["pts"] > 0 else warn_font
            lc = ws.cell(row=row, column=len(cols) - 1)
            lc.hyperlink = tv_url(t)
            lc.value = "chart"
            lc.font = link_font
            ws.cell(row=row, column=2).font = Font(bold=True)
        style_sheet(ws, len(cols), [6, 9, 26, 16, 7, 7, 30, 10, 9] + [9] * len(extra_cols)
                    + [10, 9, 8, 10, 50])

    wsA = wb.active
    wsA.title = "A. VCP"
    list_sheet(wsA, vcp[0], VCP_TIERS, VCP_ONLINE, vcp[2], vcp[1])
    wsB = wb.create_sheet("B. Weinstein 2A")
    list_sheet(wsB, stage[0], STAGE_TIERS, STAGE_ONLINE, stage[2], stage[1])
    wsC = wb.create_sheet("C. Pre-breakout")
    list_sheet(wsC, pre[0], VCP_TIERS, VCP_ONLINE, pre[2], pre[1])

    for pid, sheet_name, snaps, tiers, online, extra_cols, pred in cap_defs:
        ws = wb.create_sheet(sheet_name)
        cap_sheet(ws, snaps, tiers, online, extra_cols, pred)

    ws4 = wb.create_sheet("4. 總表")
    ws4.append(["代號", "名稱", "產業", "交易所", "上升分數", "催化", "主因", "VCP", "Weinstein",
                "Pre-breakout", "線上數", "收盤", "距高%", "市值($B)", "級距", "確定性總分"]
               + [f"{name}({w})" for _, name, w, _ in C7_COLS] + ["HL結構", "TradingView"])
    for e in sorted(rec.values(), key=lambda e: (-e.get("rise_score", 0), -e["online_count"], -e["best_score"])):
        t = e["ticker"]
        c = CERT.get(t)
        cat = CATALYST.get(t)
        mc = e.get("mcap") or 0
        cap = "大" if mc >= 10e9 else "中" if mc >= 2e9 else "小" if mc > 0 else ""
        ws4.append([t, (e["name"] or "")[:40], (e.get("sector") or "")[:24],
                    EXCHANGE.get(t, "").upper(), e.get("rise_score", 0),
                    (("🔥 " if cat["pts"] > 0 else "⚠ ") + cat["reason"]) if cat else "",
                    e.get("rise_why", ""), e["vcp"] or "", e["stage"] or "", e["pre"] or "",
                    e["online_count"], e["price"], e["off_high_pct"],
                    round(mc / 1e9, 2) or None, cap, c["cert"] if c else None]
                   + [c["c7"][k] if c else None for k, _, _, _ in C7_COLS]
                   + [("是" if c["hl_ok"] else "否") if c else "", tv_url(t)])
        row = ws4.max_row
        for col, key, onset in ((8, "vcp", VCP_ONLINE), (9, "stage", STAGE_ONLINE), (10, "pre", VCP_ONLINE)):
            cc = ws4.cell(row=row, column=col)
            cc.alignment = Alignment(horizontal="center")
            if e[key]:
                cc.fill = on_fill if e[key] in onset else off_fill
                cc.font = Font(bold=True)
        if cat:
            ws4.cell(row=row, column=6).font = hot_font if cat["pts"] > 0 else warn_font
        lc = ws4.cell(row=row, column=16 + len(C7_COLS) + 2)
        lc.hyperlink = tv_url(t)
        lc.value = "chart"
        lc.font = link_font
        ws4.cell(row=row, column=1).font = Font(bold=True)
    style_sheet(ws4, 16 + len(C7_COLS) + 2,
                [9, 26, 16, 9, 9, 28, 30, 7, 10, 12, 8, 10, 9, 10, 6, 10] + [8] * len(C7_COLS) + [8, 10])

    ws0 = wb.create_sheet("說明", 0)
    for line in [
        [f"Combined Watchlist {rev}"],
        [f"產生時間：{stamp_txt}｜模型：{model}｜數據基準：2026-08-28 官方收盤"],
        [],
        ["工作表", "內容"],
        ["A / B / C", "三份完整清單（VCP、Weinstein 2A、Pre-breakout），含各期軌跡欄。"],
        ["1a/1b/1c", "VCP TOP 50 大型（≥$10B）／中型（$2–10B）／小型（<$2B），按 VCP 分數排序。"],
        ["2a/2b/2c", "Weinstein 2A TOP 50，同樣按市值分三個級距。"],
        ["3a/3b/3c", "Pre-breakout TOP 50，按市值分三個級距。"],
        ["4. 總表", "全部代號 × 三榜等級 × 上升分數 × 確定性證據 7 項量化（各為獨立欄，可用篩選排序）。"],
        [],
        ["確定性 7 項", "突破25%＋回升10%＋守底15%＋量縮15%＋收縮10%＋RS10%＋均線15%（每項 0–100，加權＝總分）"],
        ["HL 結構", "「是」＝近 45 日存在一底高於一底結構；「否」＝突破／回升／守底／收縮 4 項記 0"],
        ["催化欄", "🔥＝本週正面新聞催化；⚠＝負面／風險事件。人工整理，隨新聞更新。"],
        ["等級底色", "綠＝該清單的線上級別；灰＝線下；空白＝未出現在該清單"],
        [],
        ["注意", "本表為技術面選股輔助，非投資建議；分級由量化規則產生，形態請開圖確認。"],
    ]:
        ws0.append(line)
    ws0.column_dimensions["A"].width = 18
    ws0.column_dimensions["B"].width = 90
    ws0["A1"].font = Font(bold=True, size=14)
    for c in ("A4", "B4"):
        ws0[c].fill = head_fill
        ws0[c].font = head_font
    wb.save(path)


# ---------------------------------------------------------------------- main
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--rev", default="R3")
    ap.add_argument("--model", default="Fable5;ultracode")
    args = ap.parse_args()

    vcp_snaps = load_vcp_snapshots()
    stage_snaps = load_series("scan_stage_R*.json", "stage")
    pre_snaps = load_series("scan_PB-R*.json", "category")

    vcp_letter = {k: v[0] for k, v in VCP_TIERS.items()}
    stage_letter = {k: v[0] for k, v in STAGE_TIERS.items()}
    vcp_tracks = build_tracks(vcp_snaps, vcp_letter)
    stage_tracks = build_tracks(stage_snaps, stage_letter)
    pre_tracks = build_tracks(pre_snaps, vcp_letter)
    vcp_ranks = rank_map(VCP_ORDER, vcp_letter)
    stage_ranks = rank_map(STAGE_ORDER, stage_letter)

    VCP_COLS = [("距高", "off_high_pct", "offhigh"), ("1月", "chg_1m", "pct"),
                ("3月", "chg_3m", "pct"), ("量比", "vol_ratio", "vol")]
    STAGE_COLS = [("距高", "off_high_pct", "offhigh"), ("6月", "chg_6m", "pct"),
                  ("1年", "chg_1y", "pct"), ("200日線上", "above_ma200", "bool")]
    PRE_COLS = [("產業", "sector", "txt"), ("距高", "off_high_pct", "offhigh"),
                ("1月", "chg_1m", "pct"), ("3月", "chg_3m", "pct"), ("量比", "vol_ratio", "vol")]
    # cap pages carry 距高 already; per-list momentum columns only
    CAP_VCP = [("1月", "chg_1m", "pct"), ("3月", "chg_3m", "pct"), ("量比", "vol_ratio", "vol")]
    CAP_STG = [("6月", "chg_6m", "pct"), ("1年", "chg_1y", "pct"), ("200日線上", "above_ma200", "bool")]
    CAP_PRE = [("1月", "chg_1m", "pct"), ("3月", "chg_3m", "pct"), ("量比", "vol_ratio", "vol")]

    all_missing = missing({r["ticker"] for s in (vcp_snaps, stage_snaps, pre_snaps)
                           for r in s[-1][3]["rows"]})
    if all_missing:
        print("WARNING unmapped exchanges:", all_missing)

    rec = build_summary(vcp_snaps, stage_snaps, pre_snaps)
    for key, snaps in (("vcp", vcp_snaps), ("stage", stage_snaps), ("pre", pre_snaps)):
        for r in snaps[-1][3]["rows"]:
            e = rec.get(r["ticker"])
            if e is not None and not e.get("mcap") and r.get("mcap"):
                e["mcap"] = r["mcap"]

    dates = collections.Counter(
        (r.get("as_of") or "")[:10]
        for s in (vcp_snaps, stage_snaps, pre_snaps) for r in s[-1][3]["rows"])
    dates.pop("", None)
    n_all = sum(dates.values())
    newest = max(dates)
    n_newest = dates[newest]
    older = n_all - n_newest
    basis = (f"{newest} 官方收盤" if older == 0
             else f"{newest} 官方收盤（{n_newest}/{n_all} 檔；其餘 {older} 檔為較早報價）")

    now_hkt = datetime.now(timezone.utc) + timedelta(hours=8)
    stamp = now_hkt.strftime("%m.%d_%H.%M")
    stamp_txt = now_hkt.strftime("%Y.%m.%d %H:%M") + " HKT"
    base = f"Combined Watchlist_{args.rev} ({args.model})_({stamp})"

    lists3 = [("1", "VCP", vcp_snaps, VCP_TIERS, VCP_ONLINE, CAP_VCP),
              ("2", "Weinstein 2A", stage_snaps, STAGE_TIERS, STAGE_ONLINE, CAP_STG),
              ("3", "Pre-breakout", pre_snaps, VCP_TIERS, VCP_ONLINE, CAP_PRE)]

    pages = [
        page_html("a", "Page A｜VCP Watchlist（含歷史軌跡）",
                  f"Minervini 波動收縮形態。追蹤 {len(vcp_snaps)} 次更新，每列內含各期級別與分數。"
                  "線上＝A（緊縮待突破）／E（已突破延伸）／B（上升結構）。",
                  vcp_snaps, vcp_tracks, VCP_TIERS, VCP_ORDER, VCP_ONLINE, vcp_ranks, VCP_COLS),
        page_html("b", "Page B｜Weinstein Stage 2A Watchlist（含歷史軌跡）",
                  "Stan Weinstein 階段分析，重點在剛脫離第一階段基底的年輕升勢。"
                  "線上＝2A（初升段）／2B（主升段）／1→2（轉強觀察）。",
                  stage_snaps, stage_tracks, STAGE_TIERS, STAGE_ORDER, STAGE_ONLINE, stage_ranks, STAGE_COLS),
        page_html("c", "Page C｜Pre-Breakout Watchlist（含歷史軌跡）",
                  "突破前候選：貼近樞紐、量縮待變的標的。線上＝A／E／B。",
                  pre_snaps, pre_tracks, VCP_TIERS, VCP_ORDER, VCP_ONLINE, vcp_ranks, PRE_COLS),
    ]
    cap_defs = []
    for no, lname, snaps, tiers, online, cols in lists3:
        for suffix, tname, tdesc, pred in CAP_TIERS:
            pid = f"{no}{suffix}"
            pages.append(cap_page(pid, no, lname, snaps, tiers, online, cols, tname, tdesc, pred))
            cap_defs.append((pid, f"{pid} {lname[:4]}{tname[:2]}", snaps, tiers, online, cols, pred))
    pages.append(summary4_html(rec))

    tab_defs = [("a", "A · VCP", f"{len(vcp_snaps[-1][3]['rows'])} 檔 · 含軌跡"),
                ("b", "B · Weinstein", f"{len(stage_snaps[-1][3]['rows'])} 檔 · 含軌跡"),
                ("c", "C · 突破前", f"{len(pre_snaps[-1][3]['rows'])} 檔 · 含軌跡")]
    for no, lname, *_ in lists3:
        for suffix, tname, _, _ in CAP_TIERS:
            short = {"1": "VCP", "2": "2A", "3": "突破前"}[no]
            tab_defs.append((f"{no}{suffix}", f"{no}{suffix} · {tname[:2]}", f"{short} TOP50"))
    tab_defs.append(("4", "4 · 總表", f"{len(rec)} 檔 · 7項確定性"))
    tabs = "".join(
        f'<button role="tab" aria-selected="{"true" if i == 0 else "false"}" '
        f'aria-controls="page-{p}" data-p="{p}">{lbl}<small>{sub}</small></button>'
        for i, (p, lbl, sub) in enumerate(tab_defs))
    page_ids = json.dumps([p for p, _, _ in tab_defs])

    doc = HTML_SHELL.format(
        rev=args.rev, model=esc(args.model), stamp=stamp_txt,
        n_tickers=len(rec), pages="\n".join(pages), basis=basis, tabs=tabs, page_ids=page_ids)
    open(f"{base}.html", "w").write(doc)

    write_excel(f"{base}.xlsx",
                (vcp_snaps, vcp_tracks, VCP_COLS),
                (stage_snaps, stage_tracks, STAGE_COLS),
                (pre_snaps, pre_tracks, PRE_COLS),
                cap_defs, rec, args.rev, args.model, stamp_txt)

    print(f"{base}.html")
    print(f"{base}.xlsx")
    print(f"tickers: {len(rec)} | vcp {len(vcp_snaps[-1][3]['rows'])} "
          f"| stage {len(stage_snaps[-1][3]['rows'])} | pre {len(pre_snaps[-1][3]['rows'])}")


HTML_SHELL = """<title>Combined Watchlist {rev}</title>
<style>
:root {{
  --bg:#F5F6F4; --panel:#FFFFFF; --ink:#1B2430; --muted:#5D6B7A; --line:#DDE2E0;
  --accent:#B07B24; --accent-soft:#F3E8D3; --up:#17835C; --dn:#C24A3F; --dry:#B07B24;
  --head:#EDEFEA; --hover:#F0F3EE; --on:#17835C; --off:#8B95A0;
  --hot:#FBEEDD; --hotink:#8A5A10; --warnbg:#F9E3E0; --warnink:#9E3A30;
  --tA:#17835C; --tE:#B07B24; --tB:#2C6E9E; --tC:#7A8794; --tD:#A8B0B8;
  --t2a:#17835C; --t2b:#2C6E9E; --t12:#B07B24; --t3:#A05A2C; --t41:#8B95A0; --tdrop:#C24A3F;
}}
@media (prefers-color-scheme: dark) {{
  :root:not([data-theme="light"]) {{
    --bg:#0E1216; --panel:#151B21; --ink:#E4E9EC; --muted:#8B98A5; --line:#26303A;
    --accent:#E5B15C; --accent-soft:#2A2418; --up:#3FB68B; --dn:#E0705F; --dry:#E5B15C;
    --head:#1A222A; --hover:#1C242D; --on:#3FB68B; --off:#5D6B7A;
    --hot:#332812; --hotink:#F0C070; --warnbg:#381F1C; --warnink:#F0958A;
    --tA:#3FB68B; --tE:#E5B15C; --tB:#5CA3D6; --tC:#6B7885; --tD:#4A555F;
    --t2a:#3FB68B; --t2b:#5CA3D6; --t12:#E5B15C; --t3:#D08A5A; --t41:#5D6B7A; --tdrop:#E0705F;
  }}
}}
:root[data-theme="dark"] {{
  --bg:#0E1216; --panel:#151B21; --ink:#E4E9EC; --muted:#8B98A5; --line:#26303A;
  --accent:#E5B15C; --accent-soft:#2A2418; --up:#3FB68B; --dn:#E0705F; --dry:#E5B15C;
  --head:#1A222A; --hover:#1C242D; --on:#3FB68B; --off:#5D6B7A;
  --hot:#332812; --hotink:#F0C070; --warnbg:#381F1C; --warnink:#F0958A;
  --tA:#3FB68B; --tE:#E5B15C; --tB:#5CA3D6; --tC:#6B7885; --tD:#4A555F;
  --t2a:#3FB68B; --t2b:#5CA3D6; --t12:#E5B15C; --t3:#D08A5A; --t41:#5D6B7A; --tdrop:#E0705F;
}}
* {{ box-sizing:border-box; }}
body {{ margin:0; background:var(--bg); color:var(--ink);
  font:15px/1.55 "Avenir Next","Segoe UI","PingFang TC","Microsoft JhengHei",system-ui,sans-serif; }}
.wrap {{ max-width:1500px; margin:0 auto; padding:26px 20px 60px; }}
.masthead {{ display:flex; flex-wrap:wrap; align-items:baseline; gap:10px 18px;
  border-bottom:3px solid var(--ink); padding-bottom:13px; }}
.masthead h1 {{ font-size:27px; margin:0; letter-spacing:-0.01em; }}
.masthead h1 em {{ font-style:normal; color:var(--accent); }}
.meta {{ color:var(--muted); font-size:13px; }} .meta b {{ color:var(--ink); font-weight:600; }}
.tabs {{ display:flex; flex-wrap:wrap; gap:4px; margin:18px 0 6px;
  border-bottom:1px solid var(--line); padding-bottom:0; }}
.tabs button {{ font:inherit; font-size:13px; font-weight:600; color:var(--muted); background:none;
  border:0; border-bottom:3px solid transparent; padding:8px 10px; cursor:pointer; border-radius:5px 5px 0 0; }}
.tabs button small {{ display:block; font-weight:400; font-size:10.5px; opacity:.75; }}
.tabs button:hover {{ background:var(--hover); color:var(--ink); }}
.tabs button[aria-selected="true"] {{ color:var(--accent); border-bottom-color:var(--accent); background:var(--panel); }}
.tabs button:focus-visible {{ outline:2px solid var(--accent); outline-offset:2px; }}
.tabs button:nth-child(3), .tabs button:nth-child(6), .tabs button:nth-child(9),
.tabs button:nth-child(12) {{ margin-right:12px; }}
.page {{ display:none; }} .page.on {{ display:block; }}
.ptitle {{ font-size:21px; margin:22px 0 4px; }}
.lede {{ margin:6px 0 0; color:var(--muted); font-size:13.5px; max-width:96ch; }}
.lede b {{ color:var(--ink); }}
h4.sec {{ font-size:14px; margin:24px 0 8px; }}
.hotbar {{ display:flex; flex-wrap:wrap; align-items:center; gap:7px; background:var(--panel);
  border:1px solid var(--line); border-left:4px solid var(--dn); border-radius:6px;
  padding:10px 14px; margin:14px 0 10px; }}
.hlabel {{ font-weight:800; font-size:13.5px; color:var(--dn); white-space:nowrap; }}
.hmacro {{ color:var(--muted); font-size:12px; flex-basis:100%; order:9; }}
.hchip {{ display:inline-block; border-radius:5px; padding:3px 9px; font-size:12px;
  text-decoration:none; white-space:nowrap; }}
.hchip b {{ font-weight:800; }}
.hchip.hot {{ background:var(--hot); color:var(--hotink); }}
.hchip.warn {{ background:var(--warnbg); color:var(--warnink); }}
.hchip:hover {{ filter:brightness(1.06); }}
.cat {{ display:inline-block; border-radius:4px; padding:1px 7px; font-size:11px; font-weight:700;
  white-space:nowrap; }}
.cat.hot {{ background:var(--hot); color:var(--hotink); }}
.cat.warn {{ background:var(--warnbg); color:var(--warnink); }}
.catcell {{ max-width:190px; }} .whycell {{ max-width:250px; white-space:normal; }}
.whycell .why {{ display:block; color:var(--muted); font-size:11.5px; line-height:1.4; }}
.captier {{ display:inline-block; margin-left:5px; color:var(--muted); font-size:10px; }}
.summary {{ display:flex; gap:9px; flex-wrap:wrap; margin:15px 0 6px; }}
.summary a {{ text-decoration:none; color:inherit; }}
.stat {{ display:flex; align-items:center; gap:8px; background:var(--panel);
  border:1px solid var(--line); border-radius:6px; padding:7px 13px; }}
.stat.tot {{ border-style:dashed; }}
.stat .dot {{ width:10px; height:10px; border-radius:2px; }}
.stat b {{ font-size:19px; font-variant-numeric:tabular-nums; }}
.stat span {{ color:var(--muted); font-size:12.5px; }}
.d-a, .d-2a {{ background:var(--tA); }} .d-e, .d-12 {{ background:var(--tE); }}
.d-b, .d-2b {{ background:var(--tB); }} .d-c, .d-3 {{ background:var(--tC); }}
.d-d, .d-41 {{ background:var(--tD); }}
.market {{ background:var(--panel); border:1px solid var(--line); border-left:4px solid var(--accent);
  border-radius:6px; padding:13px 17px; margin:14px 0 4px; font-size:13.5px; color:var(--muted); }}
.market strong {{ color:var(--ink); }}
.tier {{ margin-top:30px; }}
.tier-head {{ display:flex; align-items:flex-start; gap:13px; margin-bottom:9px; }}
.tier-head h3 {{ margin:0; font-size:17px; }}
.tier-head p {{ margin:2px 0 0; color:var(--muted); font-size:13px; max-width:74ch; }}
.tier-head .count {{ margin-left:auto; color:var(--muted); font-size:13px; white-space:nowrap; padding-top:3px; }}
.badge {{ flex:none; min-width:32px; height:32px; border-radius:6px; display:grid; place-items:center;
  font-weight:700; font-size:14px; color:var(--bg); padding:0 5px; }}
.b-a, .b-2a {{ background:var(--tA); }} .b-e, .b-12 {{ background:var(--tE); }}
.b-b, .b-2b {{ background:var(--tB); }} .b-c, .b-3 {{ background:var(--tC); }}
.b-d, .b-41 {{ background:var(--tD); }} .b-drop {{ background:var(--tdrop); }}
.divider {{ display:flex; align-items:center; gap:13px; margin:40px 0 0; }}
.divider hr {{ flex:1; border:0; border-top:2px dashed var(--accent); margin:0; }}
.divider span {{ color:var(--accent); font-size:12.5px; font-weight:700; letter-spacing:0.08em; }}
.tblwrap {{ overflow-x:auto; border:1px solid var(--line); border-radius:8px; background:var(--panel); }}
table {{ border-collapse:collapse; width:100%; font-size:13.5px; }}
th {{ background:var(--head); text-align:left; padding:7px 9px; font-size:11.5px; text-transform:uppercase;
  letter-spacing:0.04em; color:var(--muted); white-space:nowrap; position:sticky; top:0; }}
th small {{ display:block; font-size:10px; letter-spacing:0; text-transform:none; font-weight:400; }}
th.sort {{ cursor:pointer; user-select:none; }}
th.sort:hover {{ color:var(--ink); background:var(--hover); }}
th.sort::after {{ content:"⇅"; opacity:.45; margin-left:4px; font-size:10px; }}
th.sort[data-dir="desc"]::after {{ content:"▼"; opacity:1; color:var(--accent); }}
th.sort[data-dir="asc"]::after {{ content:"▲"; opacity:1; color:var(--accent); }}
th.c7, td.c7 {{ border-left:1px dotted var(--line); }}
th.certsum, td.certsum {{ border-left:2px solid var(--accent); }}
td.certsum small {{ display:block; color:var(--muted); font-size:9.5px; }}
.sb2 i {{ background:var(--tB) !important; }}
td {{ padding:7px 9px; border-top:1px solid var(--line); vertical-align:top; }}
tbody tr:hover {{ background:var(--hover); }}
td.num, th.num {{ text-align:right; font-variant-numeric:tabular-nums;
  font-family:"SF Mono","Cascadia Mono",Consolas,ui-monospace,monospace; font-size:12.5px; white-space:nowrap; }}
td.tk {{ white-space:nowrap; }}
td.tk a {{ display:inline-block; font-weight:700; color:var(--accent); text-decoration:none;
  border-bottom:1px solid transparent; }}
td.tk a small {{ display:block; font-size:9.5px; font-weight:600; letter-spacing:0.06em; color:var(--muted); }}
td.tk a:hover, td.tk a:focus-visible {{ border-bottom-color:var(--accent); outline:none; }}
td.nm {{ color:var(--muted); font-size:12.5px; white-space:nowrap; }}
th.st, td.st, th.gr, td.gr {{ text-align:center; }}
td.st {{ white-space:nowrap; font-variant-numeric:tabular-nums; }}
td.st b {{ display:inline-block; min-width:19px; height:19px; line-height:19px; border-radius:4px;
  color:var(--bg); font-size:11px; padding:0 3px; }}
td.st i {{ display:block; font-style:normal; font-size:10px; color:var(--muted); margin-top:1px; }}
td.st.none, td.gr.none {{ color:var(--muted); }}
td.tA b, td.t2A b, td.t2a b {{ background:var(--tA); }} td.tE b, td.te b {{ background:var(--tE); }}
td.tB b, td.t2B b, td.tb b, td.t2b b {{ background:var(--tB); }}
td.tC b, td.t3 b, td.tc b {{ background:var(--tC); }}
td.tD b, td.t41 b, td.td b {{ background:var(--tD); }} td.t12 b {{ background:var(--t12); }}
td.ta b {{ background:var(--tA); }}
td.gr b {{ display:inline-block; min-width:26px; padding:1px 6px; border-radius:4px;
  font-size:11.5px; color:var(--bg); }}
td.gr.on b {{ background:var(--on); }} td.gr.off b {{ background:var(--off); }}
tr.cur td {{ background:var(--accent-soft); font-weight:600; }}
td.tr, th.tr {{ text-align:center; }}
.mv {{ font-size:13px; }} .mv.up {{ color:var(--up); }} .mv.dn {{ color:var(--dn); }} .mv.fl {{ color:var(--muted); }}
td.up {{ color:var(--up); }} td.dn {{ color:var(--dn); }}
td.pivot {{ color:var(--accent); font-weight:600; }} td.dry {{ color:var(--dry); font-weight:600; }}
td.note {{ min-width:150px; max-width:300px; color:var(--muted); font-size:12px; line-height:1.45; white-space:normal; }}
.chip-cal {{ display:inline-block; background:var(--accent-soft); color:var(--accent); border-radius:4px;
  padding:1px 6px; font-size:11px; font-weight:600; white-space:nowrap; }}
.scorebar {{ display:inline-flex; align-items:center; gap:6px; min-width:84px; }}
.scorebar i {{ display:block; height:5px; border-radius:3px; background:var(--accent);
  min-width:3px; max-width:52px; flex:none; }}
.scorebar b {{ font-size:12.5px; }}
.dwrap {{ display:flex; flex-wrap:wrap; gap:6px; }}
.dchip {{ display:inline-flex; flex-direction:column; background:var(--panel); border:1px solid var(--line);
  border-radius:5px; padding:4px 9px; text-decoration:none; color:var(--muted); font-size:12.5px; font-weight:700; }}
.dchip small {{ font-weight:400; font-size:10px; letter-spacing:0.04em; }}
.dchip:hover, .dchip:focus-visible {{ border-color:var(--accent); color:var(--ink); outline:none; }}
.method {{ margin-top:40px; border-top:1px solid var(--line); padding-top:18px;
  color:var(--muted); font-size:13px; max-width:96ch; }}
.method h3 {{ color:var(--ink); font-size:15px; margin:0 0 8px; }}
.method ul {{ padding-left:20px; }} .method li {{ margin:4px 0; }}
.disclaimer {{ margin-top:13px; font-size:12.5px; border-left:3px solid var(--dn); padding-left:12px; }}
a {{ color:var(--accent); }}
@media (max-width:640px) {{ .masthead h1 {{ font-size:22px; }} .tier-head .count {{ display:none; }} }}
</style>
<div class="wrap">
<header class="masthead">
  <h1>Combined Watchlist <em>{rev}</em></h1>
  <span class="meta"><b>VCP · Weinstein 2A · Pre-breakout · 市值分級 · 確定性 7 項</b>｜合計 {n_tickers} 檔｜
  數據基準 {basis}｜產生 {stamp}｜{model}</span>
</header>
<nav class="tabs" role="tablist">{tabs}</nav>
{pages}
<footer class="method">
<h3>使用說明</h3>
<ul>
<li><b>Page A／B／C</b>：三份完整清單的最新版本，含歷次掃描軌跡（有色方塊＝該期在線上）；▲ 升級、▼ 降級。</li>
<li><b>Page 1a–3c</b>：每份清單依官方市值分三個級距（大 ≥$10B／中 $2–10B／小 <$2B），各取分數 TOP 50。頁頂紅框
「本週熱點催化」集中列出該頁涉及的新聞驅動事件（🔥 正面／⚠ 風險），表內催化欄逐檔標示。</li>
<li><b>Page 4 總表</b>：所有代號 × 三榜等級 × 上升就緒分數 × 市值 × <b>確定性證據 7 項量化</b>（突破／回升／守底／
量縮／收縮／RS／均線，各 0–100 獨立成欄）。<b>點任何數值欄標題即依該欄重新排序</b>（第一下降序、再點升序）。</li>
<li>確定性 7 項取自 10MA 上升趨勢清單的算法：以官方日線序列偵測「一底高於一底」結構後計算突破進度、跌幅收復、
低點守住天數、量能對比、回檔收縮、相對強度、均線排列；加權合計＝確定性總分（權重 25/10/15/15/10/10/15）。</li>
<li>數據源升級：本版全部收盤價、市值與 1 月／3 月動能改用官方每日快照序列（240/274 檔全序列；
CRWD 4:1、KLAC 10:1、DD 1:3 拆股已調整），其餘 34 檔（多為外國 ADR）採官方 8/28 單日收盤。</li>
<li>點任一代號可開 TradingView 圖表（Weinstein 建議切週線＋30 週均線）。</li>
</ul>
<p class="disclaimer">本表為技術面選股輔助工具，非投資建議。分級由量化規則產生，催化欄為人工整理的本週事件，
形態最終以圖表確認為準。</p>
</footer>
</div>
<script>
(function () {{
  var tabs = Array.prototype.slice.call(document.querySelectorAll('.tabs button'));
  var pages = Array.prototype.slice.call(document.querySelectorAll('.page'));
  var ids = {page_ids};
  function show(id, push) {{
    tabs.forEach(function (b) {{ b.setAttribute('aria-selected', b.dataset.p === id ? 'true' : 'false'); }});
    pages.forEach(function (p) {{ p.classList.toggle('on', p.id === 'page-' + id); }});
    try {{ localStorage.setItem('cw3-page', id); }} catch (e) {{}}
    if (push) {{ try {{ history.replaceState(null, '', '#page-' + id); }} catch (e) {{}} }}
  }}
  tabs.forEach(function (b) {{
    b.addEventListener('click', function () {{ show(b.dataset.p, true); window.scrollTo({{top: 0}}); }});
    b.addEventListener('keydown', function (e) {{
      var i = tabs.indexOf(b), n = null;
      if (e.key === 'ArrowRight') n = tabs[(i + 1) % tabs.length];
      if (e.key === 'ArrowLeft') n = tabs[(i - 1 + tabs.length) % tabs.length];
      if (n) {{ e.preventDefault(); n.focus(); show(n.dataset.p, true); }}
    }});
  }});
  var start = 'a';
  var h = (location.hash || '').replace('#page-', '');
  if (ids.indexOf(h) >= 0) start = h;
  else {{ try {{ var s = localStorage.getItem('cw3-page'); if (s && ids.indexOf(s) >= 0) start = s; }} catch (e) {{}} }}
  show(start, false);
  document.addEventListener('click', function (e) {{
    var a = e.target.closest && e.target.closest('a[href^="#"]');
    if (!a) return;
    var t = document.querySelector(a.getAttribute('href'));
    if (!t) return;
    var pg = t.closest('.page');
    if (pg && !pg.classList.contains('on')) show(pg.id.replace('page-', ''), true);
  }});
  // column sorting: first click = descending, second = ascending
  document.querySelectorAll('table.sortable').forEach(function (tb) {{
    var ths = Array.prototype.slice.call(tb.querySelectorAll('thead th'));
    ths.forEach(function (th, ci) {{
      if (!th.classList.contains('sort')) return;
      th.addEventListener('click', function () {{
        var dir = th.dataset.dir === 'desc' ? 'asc' : 'desc';
        ths.forEach(function (o) {{ delete o.dataset.dir; }});
        th.dataset.dir = dir;
        var tbody = tb.tBodies[0];
        var rows = Array.prototype.slice.call(tbody.rows);
        rows.sort(function (a, b) {{
          var av = parseFloat(a.cells[ci].getAttribute('data-v'));
          var bv = parseFloat(b.cells[ci].getAttribute('data-v'));
          if (isNaN(av)) av = -1e18;
          if (isNaN(bv)) bv = -1e18;
          return dir === 'desc' ? bv - av : av - bv;
        }});
        rows.forEach(function (r) {{ tbody.appendChild(r); }});
      }});
    }});
  }});
}})();
</script>"""


if __name__ == "__main__":
    main()
